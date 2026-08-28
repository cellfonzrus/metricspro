"""Carrier Reconciliation — PURE workbook parser (no I/O).

The back office pivots the incoming (rebate/commission/ePay) + outgoing (device cost) feeds into a
"Rebate Reconciliation" workbook every period. This module turns that workbook's bytes into a normalized
dict so MetricsPro can show, per store, the BACK-OFFICE (Boost/ePay) figure beside OUR computed figure.

DISPLAY / ANALYSIS ONLY. Nothing here reads or writes a money table, and nothing recomputes pay — it is a
byte-in / dict-out reader. The comparison against OUR engines happens in the router; this file never
touches the database.

Carrier-generic by shape: the workbook layout (a Reports deliverable with stacked rebate blocks + a
parallel commissions block, plus Escalation / Unpaid / Missing / raw feed sheets) is the Boost/ePay
back office's format. Total/VidaPay can reuse the same normalized shape later.

The Reports sheet is the deliverable. In cols A–I it stacks THREE blocks, each starting with a
`Stores | Total Device Cost | …` (or reimbursement) header row and ending in a `Grand Total` row:
  • Block 1  — per-STORE rebate summary (the authoritative store totals we read).
  • Block 2  — per-store reimbursement breakdown (New Act / PIC / Upgrade / … ). Skipped for totals.
  • Block 3  — per-store rows WITH nested per-REP rows underneath (the rep drilldown).
In cols L–P it carries a parallel per-STORE commissions block aligned with block 1's rows:
  `Stores | Total Comm Paid | Estimated GP | Withhold | ePay Paid`.

Header rows are found BY LABEL, never by fixed row number, so a shifted export still parses. Rep rows in
block 3 are identified because their first cell is NOT one of the store names block 1 established (they are
`Last, First` people, not addresses)."""

from __future__ import annotations

import datetime as _dt
import io
from typing import Any

import openpyxl


# ── small pure helpers ─────────────────────────────────────────────────────────────────────────────
def _norm(v: Any) -> str:
    """Lower-cased, whitespace-collapsed label for header/keyword matching."""
    return " ".join(str(v or "").strip().split()).lower()


def _s(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _num(v: Any) -> float:
    """Parse a workbook cell to float; blanks / strings / errors → 0.0 (never raises)."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "#REF!", "#N/A", "#VALUE!"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return 0.0


def _cell(v: Any) -> Any:
    """JSON-safe scalar for display rows: datetimes → ISO strings, everything else passed through."""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return v


def _r2(x: float) -> float:
    return round(x, 2)


# ── header label constants ─────────────────────────────────────────────────────────────────────────
_GRAND_TOTAL = "grand total"

# Block 1 / Block 3 rebate-summary columns (1-based worksheet columns A–I).
_REB_COLS = {
    "store": 1, "device_cost": 2, "cust_paid": 3, "financed": 4,
    "rebate_expected": 5, "rebate_paid": 6, "rebate_diff": 7,
    "rebate_overpaid": 8, "gp": 9,
}
# Parallel commissions block columns (L–P).
_COMM_COLS = {"store": 12, "comm_paid": 13, "comm_gp": 14, "withhold": 15, "epay_paid": 16}

# Unpaid Devices — the curated display subset (the sheet carries 60+ POS columns).
_UNPAID_FIELDS = [
    "Store", "Sales Rep", "Trans Date Time", "Trans ID", "Act ID", "Terminal ID", "Device",
    "Cost", "SRP", "Disc.", "Ext Price", "Phone No", "IMEI", "ESN", "SKU", "Contract Type",
]


def _is_rebate_header(row: tuple) -> bool:
    """A `Stores | Total Device Cost | …` header row (block 1 or block 3)."""
    return (len(row) >= 2 and _norm(row[0]) == "stores"
            and _norm(row[1]) == "total device cost")


def _reb_store_summary(row: tuple) -> dict:
    """Read the rebate-summary figures (cols A–I) off one worksheet row (0-based tuple)."""
    return {k: _num(row[c - 1]) if k != "store" else _s(row[c - 1])
            for k, c in _REB_COLS.items()}


def _comm_store_summary(row: tuple) -> dict:
    """Read the commissions figures (cols L–P) off one worksheet row (0-based tuple)."""
    if len(row) < max(_COMM_COLS.values()):
        return {}
    return {k: _num(row[c - 1]) if k != "store" else _s(row[c - 1])
            for k, c in _COMM_COLS.items()}


# ── the Reports sheet ────────────────────────────────────────────────────────────────────────────
def _parse_reports(ws) -> dict:
    """Parse cols A–I (three stacked rebate blocks) + cols L–P (commissions) of the Reports sheet."""
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]

    # Locate the rebate-summary header rows by label. The FIRST is block 1 (authoritative store totals);
    # the LAST is block 3 (store rows with nested reps). If only one exists, block 3 == block 1.
    header_idxs = [i for i, r in enumerate(rows) if _is_rebate_header(r)]
    if not header_idxs:
        return {"stores": [], "totals": {}}
    b1_hdr = header_idxs[0]
    b3_hdr = header_idxs[-1]

    stores: list[dict] = []
    by_name: dict[str, dict] = {}
    totals: dict[str, float] = {}

    # ── Block 1: per-STORE rebate summary (cols A–I) + parallel commissions (cols L–P) ───────────────
    for r in rows[b1_hdr + 1:]:
        name = _s(r[0])
        if not name:
            continue
        if _norm(name) == _GRAND_TOTAL:
            gt = _reb_store_summary(r)
            for k, v in gt.items():
                if k != "store":
                    totals[k] = _r2(v)
            comm_gt = _comm_store_summary(r)   # same row carries the commission Grand Total in L–P
            for k in ("comm_paid", "comm_gp", "withhold", "epay_paid"):
                totals[k] = _r2(comm_gt.get(k, 0.0))
            break
        rec = _reb_store_summary(r)
        rec["reps"] = []
        for k in ("comm_paid", "comm_gp", "withhold", "epay_paid"):
            rec[k] = 0.0
        stores.append(rec)
        by_name[_norm(name)] = rec

    store_names = set(by_name.keys())

    # Commissions block (cols L–P) — aligned with block 1's store rows, keyed by store NAME so a row
    # offset never mis-attributes money. Stop at its own Grand Total.
    for r in rows[b1_hdr + 1:]:
        cs = _comm_store_summary(r)
        cname = _norm(cs.get("store"))
        if not cname:
            continue
        if cname == _GRAND_TOTAL:
            break
        tgt = by_name.get(cname)
        if tgt:
            tgt["comm_paid"] = cs["comm_paid"]
            tgt["comm_gp"] = cs["comm_gp"]
            tgt["withhold"] = cs["withhold"]
            tgt["epay_paid"] = cs["epay_paid"]

    # ── Block 3: rep drilldown. A first-cell that IS a known store opens that store; anything else is a
    # rep row (`Last, First`) belonging to the current store. Rep rebate figures come from cols A–I;
    # the commission cols in block 3 are NOT row-aligned to the rebate cols, so they are left out of the
    # rep rebate figures. ───────────────────────────────────────────────────────────────────────────
    if b3_hdr != b1_hdr:
        current: dict | None = None
        for r in rows[b3_hdr + 1:]:
            name = _s(r[0])
            if not name:
                continue
            if _norm(name) == _GRAND_TOTAL:
                break
            key = _norm(name)
            if key in store_names:
                current = by_name.get(key)
                continue
            if current is not None:
                rep = _reb_store_summary(r)
                rep["rep"] = rep.pop("store")
                current["reps"].append(rep)

    totals["store_count"] = len(stores)
    return {"stores": stores, "totals": totals}


# ── line-item sheets ─────────────────────────────────────────────────────────────────────────────
def _parse_line_sheet(ws, fields: list[str] | None = None) -> tuple[list[dict], list[str]]:
    """Generic header-row-1 line-item reader. `fields` optionally restricts to a curated column subset
    (matched by header label). Returns (rows, headers_kept)."""
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return [], []
    hdr = [_s(h) for h in header]
    if fields:
        want = {_norm(f): f for f in fields}
        idxs = [(i, hdr[i]) for i in range(len(hdr)) if _norm(hdr[i]) in want]
    else:
        idxs = [(i, hdr[i]) for i in range(len(hdr)) if hdr[i]]
    kept = [h for _, h in idxs]
    out: list[dict] = []
    for row in it:
        if row is None:
            continue
        if all(c is None or _s(c) == "" for c in row):
            continue
        out.append({h: _cell(row[i]) if i < len(row) else None for i, h in idxs})
    return out, kept


def _parse_escalations(ws) -> tuple[list[dict], float]:
    rows, _ = _parse_line_sheet(ws)
    # Expected Amount total — the escalation headline the pivot ("Escalation Report") also reports.
    exp_key = next((k for k in (rows[0].keys() if rows else []) if _norm(k) == "expected amount"), None)
    total = _r2(sum(_num(r.get(exp_key)) for r in rows)) if exp_key else 0.0
    return rows, total


def _parse_unpaid(ws) -> tuple[list[dict], float]:
    rows, _ = _parse_line_sheet(ws, _UNPAID_FIELDS)
    cost_key = next((k for k in (rows[0].keys() if rows else []) if _norm(k) == "cost"), None)
    total = _r2(sum(_num(r.get(cost_key)) for r in rows)) if cost_key else 0.0
    return rows, total


def _count_data_rows(ws) -> int:
    """Count non-blank data rows below the header (row 1) — the raw per-transaction feed size."""
    it = ws.iter_rows(values_only=True)
    try:
        next(it)   # header
    except StopIteration:
        return 0
    n = 0
    for row in it:
        if row is None:
            continue
        if all(c is None or _s(c) == "" for c in row):
            continue
        n += 1
    return n


def _find_sheet(wb, *names: str):
    """Case/space-insensitive sheet lookup; None if absent."""
    want = {_norm(n) for n in names}
    for ws in wb.worksheets:
        if _norm(ws.title) in want:
            return ws
    return None


# ── public entry point ─────────────────────────────────────────────────────────────────────────────
def parse_workbook(xlsx_bytes: bytes, *, period: str | None = None) -> dict:
    """Parse a Carrier Reconciliation workbook (bytes) into the normalized structure.

    PURE: bytes in, dict out — no database, no recompute. Returns:
      {
        period,
        stores: [ {store, device_cost, cust_paid, financed, rebate_expected, rebate_paid, rebate_diff,
                   rebate_overpaid, gp, comm_paid, comm_gp, withhold, epay_paid, reps:[…]} ],
        escalations: [ {…display row…} ],
        unpaid_devices: [ {…curated POS columns…} ],
        missing: [ {…display row…} ],
        totals: { device_cost, cust_paid, financed, rebate_expected, rebate_paid, rebate_diff,
                  rebate_overpaid, gp, comm_paid, comm_gp, withhold, epay_paid, store_count,
                  escalation_count, escalation_expected, unpaid_count, unpaid_cost, missing_count },
        raw_txn_count,
      }
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    # Reports sheet (deliverable). "Report" (with #REF! headers) is the superseded variant — ignore it.
    reports_ws = _find_sheet(wb, "Reports")
    parsed = _parse_reports(reports_ws) if reports_ws else {"stores": [], "totals": {}}
    stores = parsed["stores"]
    totals = dict(parsed["totals"])

    esc_ws = _find_sheet(wb, "Escalation")
    escalations, esc_total = _parse_escalations(esc_ws) if esc_ws else ([], 0.0)

    unpaid_ws = _find_sheet(wb, "Unpaid Devices", "Unpaid")
    unpaid, unpaid_cost = _parse_unpaid(unpaid_ws) if unpaid_ws else ([], 0.0)

    missing_ws = _find_sheet(wb, "Missing")
    missing, _ = _parse_line_sheet(missing_ws) if missing_ws else ([], [])

    raw_ws = _find_sheet(wb, "Sheet1")
    raw_txn_count = _count_data_rows(raw_ws) if raw_ws else 0

    try:
        wb.close()
    except Exception:
        pass

    totals["escalation_count"] = len(escalations)
    totals["escalation_expected"] = esc_total
    totals["unpaid_count"] = len(unpaid)
    totals["unpaid_cost"] = unpaid_cost
    totals["missing_count"] = len(missing)
    totals.setdefault("store_count", len(stores))

    return {
        "period": period,
        "stores": stores,
        "escalations": escalations,
        "unpaid_devices": unpaid,
        "missing": missing,
        "totals": totals,
        "raw_txn_count": raw_txn_count,
    }


# ── Boost-vs-ours comparison (PURE merge/diff) ──────────────────────────────────────────────────────
# The dollar fields compared side by side. `rebate_expected` is Boost-only context (we have no
# expectation feed); the diff is computed only on the four figures we compute on OUR side.
_COMPARE_FIELDS = ("rebate_paid", "comm_paid", "epay_paid", "gp")
_BOOST_FIELDS = ("rebate_paid", "rebate_expected", "comm_paid", "epay_paid", "gp",
                 "device_cost", "rebate_diff")


def build_comparison(parsed: dict, ours_by_store: dict[str, dict], *,
                     resolve=None, tol: float = 0.01) -> dict:
    """Merge the parsed Boost/back-office per-store figures with OUR computed figures and diff them.

    PURE — no I/O. The caller supplies:
      • `parsed`         — the `parse_workbook` result.
      • `ours_by_store`  — `{normalized_store_key: {rebate_paid, comm_paid, epay_paid, gp}}`, OUR figures
                           already keyed on the canonical store key.
      • `resolve`        — optional `workbook_store_name -> canonical_label`. Its output (lower-cased) is
                           the first key tried against `ours_by_store`; the raw workbook name (lower-cased)
                           is the fallback. `None` = identity (the raw name is the only key).

    A store is `matched` when an OUR row is found for it. A store that matches nothing is STILL returned
    (with zeroed `ours` and a full `diff`) AND named in `unmatched_stores`, so a workbook store is never
    silently dropped. `diff = boost - ours`, rounded to the cent; `match_ok` per field flags |diff| ≤ tol.

    Returns `{period, per_store:[…], unmatched_stores:[…], totals:{boost, ours, diff}}`."""
    per_store: list[dict] = []
    unmatched: list[str] = []
    b_tot = {f: 0.0 for f in _COMPARE_FIELDS}
    o_tot = {f: 0.0 for f in _COMPARE_FIELDS}

    for s in parsed.get("stores", []):
        name = s.get("store") or ""
        resolved = None
        if resolve is not None:
            try:
                resolved = resolve(name)
            except Exception:
                resolved = None
        o = None
        if resolved:
            o = ours_by_store.get(_norm(resolved))
        if o is None:
            o = ours_by_store.get(_norm(name))
        matched = o is not None
        o = o or {}

        boost = {f: _r2(_num(s.get(f))) for f in _BOOST_FIELDS}
        ours = {f: _r2(_num(o.get(f))) for f in _COMPARE_FIELDS}
        diff = {f: _r2(boost[f] - ours[f]) for f in _COMPARE_FIELDS}
        match_ok = {f: abs(diff[f]) <= tol for f in _COMPARE_FIELDS}

        for f in _COMPARE_FIELDS:
            b_tot[f] = _r2(b_tot[f] + boost[f])
            o_tot[f] = _r2(o_tot[f] + ours[f])

        per_store.append({
            "store": name,
            "resolved_store": resolved,
            "matched": matched,
            "boost": boost,
            "ours": ours,
            "diff": diff,
            "match_ok": match_ok,
        })
        if not matched:
            unmatched.append(name)

    per_store.sort(key=lambda r: r["store"].lower())
    diff_tot = {f: _r2(b_tot[f] - o_tot[f]) for f in _COMPARE_FIELDS}
    return {
        "period": parsed.get("period"),
        "per_store": per_store,
        "unmatched_stores": unmatched,
        "totals": {"boost": b_tot, "ours": o_tot, "diff": diff_tot},
    }
