"""Vendor Rebate / Commission History report (RQ / Wireless Zone → WirelessDotCom / Verizon).

A tenant uploads the carrier's rebate/commission history export. Each row is an activation the carrier
paid on; the dollar figure (`Unit Rebate`) is EITHER a device rebate OR a rate-plan/activation
commission, and WHICH it is, is encoded in the `Product Name`. From one file we produce three things:

  1. CUSTOMERS   — name + phone + ZIP per line (deduped by phone on write).
  2. ACTIVATIONS — device (Related Product + IMEI), rate plan, sold-on date, salesperson, contract #.
  3. P&L / GROSS PROFIT — commission → `carrier_comm`; device cost → `device_cost` (COGS); device
     rebate → `device_rebate` (contra-COGS). Net device COGS = cost − rebate, so device GP = rebate − cost.

OWNER P&L RULE (2026-08 — nothing hardcoded, no margin constant):
    device gross profit = total rebate − COGS (i.e. rebate − Related Cost)
  For a receipt-only tenant (MetricsPro is the SECONDARY POS) the sales feed does NOT book these
  financed devices — receipt-import sales write no sale_items, so neither device cost nor down payment
  reaches the P&L (verified 2026-08). This report is therefore the source of the device COST as well
  as the rebate, so BOTH are posted here: cost → device_cost (COGS), rebate → device_rebate
  (contra-COGS). Booking only the rebate (as an earlier draft did) would show the full rebate as
  phantom gross profit. The customer DOWN PAYMENT is a financing pass-through (it reduces the
  customer's financed balance owed to the carrier), NOT store device margin — it belongs on the
  BALANCE SHEET (total received), not P&L GP, so it is deliberately not posted to the P&L here.
  BYOD lines carry no device cost/rebate (commission only); an upfront purchase is price − cost on the
  sales side. The `device_gp` field below is `rebate − cost` = the booked device gross profit.

Owner-confirmed classification (2026-08):
  • "Device Payment Agreement Rebate Amount"  → DEVICE REBATE   (GP = Unit Rebate − Related Cost)
  • every other rate-plan / activation / upgrade line (New Activation, DPA New Act iPhone/iPad,
    DPA Upgrade iPhone, Upgrade - ISPU, NEW - ISPU)  → COMMISSION income
  • "Verizon Carrier Trade In Rebate" / any 'trade in' line → PASS-THROUGH (no P&L income)
  • a blank Product Name row is a report TOTALS row → dropped

This module is split I/O (read_xlsx, openpyxl) vs PURE (classify / normalize_report), so the mapping +
classification + aggregation are unit-testable without a file — see backend/harness_vendor_rebate_report.py.
Nothing writes here; the router turns the normalized result into customers/activations and the P&L feed.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

# ── classification (owner rules) ──────────────────────────────────────────────────────────────────
KIND_DEVICE_REBATE = "device_rebate"
KIND_COMMISSION = "commission"
KIND_TRADE_IN = "trade_in"
KIND_SKIP = "skip"

_DEVICE_REBATE_NAME = "device payment agreement rebate amount"


def classify(product_name: Any) -> str:
    s = str(product_name or "").strip().lower()
    if not s:
        return KIND_SKIP                       # blank Product Name = the report's totals row
    if "trade in" in s or "trade-in" in s:
        return KIND_TRADE_IN                    # pass-through, no P&L income
    if s == _DEVICE_REBATE_NAME:
        return KIND_DEVICE_REBATE               # GP = rebate − Related Cost
    return KIND_COMMISSION                       # every other rate-plan/activation/upgrade line


# ── value coercion (pure) ─────────────────────────────────────────────────────────────────────────
def money(v: Any):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if s in ("", "-", ".", "-."):
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _digits(v: Any):
    if v is None:
        return None
    d = re.sub(r"\D", "", str(v))
    return d or None


def _phone(v: Any):
    """A 10–11 digit US phone (the activated line's number). Rejects alphanumerics (e.g. a trade-in
    reference like 'E254401434291') and short/long junk."""
    d = _digits(v)
    if not d:
        return None
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    return d if len(d) == 10 else None


def _iso_date(v: Any):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def _period(iso: str | None):
    return iso[:7] if iso and len(iso) >= 7 else None   # 'YYYY-MM'


def _split_name(full: Any):
    parts = str(full or "").replace(",", " ").split()
    if not parts:
        return ("", "")
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], " ".join(parts[1:]))


# Header → canonical key. Matched case-insensitively so a small header rename still binds.
_COLS = {
    "invoice #": "invoice_no", "tracking #": "tracking_no", "qty": "qty",
    "product sku": "rebate_sku", "product name": "product_name",
    "unit rebate": "unit_rebate", "total rebate": "total_rebate",
    "related product": "device_name", "related sku": "device_sku", "related sn": "imei",
    "related cost": "related_cost", "related price": "related_price",
    "rate plan": "rate_plan", "term code": "term_code",
    "customer": "customer", "sales person": "salesperson", "sold on": "sold_on",
    "invoiced at": "store_name", "contract #": "contract_no",
    "customer identifier": "customer_identifier", "zip code": "zip",
    "region": "region", "district": "district", "vendor account name": "vendor_account",
}


def normalize_report(rows: list[dict]) -> dict:
    """PURE: normalize the report rows (each a dict keyed by canonical column) into activation records
    + a store/period P&L summary. Deterministic; no I/O."""
    activations: list[dict] = []
    dropped = {"totals_row": 0, "trade_in": 0, "no_amount": 0}
    # summary[(store, period)] = accumulator
    summ: dict[tuple, dict] = {}

    for r in rows:
        pn = r.get("product_name")
        kind = classify(pn)
        if kind == KIND_SKIP:
            dropped["totals_row"] += 1
            continue

        unit_rebate = money(r.get("unit_rebate")) or 0.0
        related_cost = money(r.get("related_cost")) or 0.0
        sold_iso = _iso_date(r.get("sold_on"))
        period = _period(sold_iso)
        store = str(r.get("store_name") or "").strip() or None
        phone = _phone(r.get("tracking_no")) or _phone(r.get("customer_identifier"))
        first, last = _split_name(r.get("customer"))

        commission = round(unit_rebate, 2) if kind == KIND_COMMISSION else 0.0
        device_rebate = round(unit_rebate, 2) if kind == KIND_DEVICE_REBATE else 0.0
        device_cost = round(related_cost, 2) if kind == KIND_DEVICE_REBATE else 0.0
        device_gp = round(device_rebate - device_cost, 2) if kind == KIND_DEVICE_REBATE else 0.0

        if kind == KIND_TRADE_IN:
            dropped["trade_in"] += 1

        activations.append({
            "invoice_no": str(r.get("invoice_no") or "").strip() or None,
            "kind": kind,
            "product_name": str(pn or "").strip() or None,
            "customer_name": str(r.get("customer") or "").strip() or None,
            "first_name": first or None, "last_name": last or None,
            "phone": phone, "zip": str(r.get("zip") or "").strip() or None,
            "device_name": str(r.get("device_name") or "").strip() or None,
            "device_sku": str(r.get("device_sku") or "").strip() or None,
            "imei": _digits(r.get("imei")),
            "rate_plan": str(r.get("rate_plan") or "").strip() or None,
            "cell_number": _phone(r.get("tracking_no")),
            "account_number": str(r.get("customer_identifier") or "").strip() or None,
            "contract_no": str(r.get("contract_no") or "").strip() or None,
            "salesperson": str(r.get("salesperson") or "").strip() or None,
            "sold_on": sold_iso, "period": period, "store_name": store,
            "carrier": "Verizon",
            "unit_rebate": round(unit_rebate, 2),
            "related_cost": round(related_cost, 2),
            "commission": commission,
            "device_rebate": device_rebate,
            "device_cost": device_cost,
            "device_gp": device_gp,
        })

        key = (store, period)
        acc = summ.setdefault(key, {
            "store_name": store, "period": period, "rows": 0,
            "commission_income": 0.0, "device_rebate": 0.0, "device_cost": 0.0,
            "device_gp": 0.0, "activations": 0,
        })
        acc["rows"] += 1
        acc["commission_income"] = round(acc["commission_income"] + commission, 2)
        acc["device_rebate"] = round(acc["device_rebate"] + device_rebate, 2)
        acc["device_cost"] = round(acc["device_cost"] + device_cost, 2)
        acc["device_gp"] = round(acc["device_gp"] + device_gp, 2)
        if kind != KIND_TRADE_IN:
            acc["activations"] += 1

    totals = {
        "rows_in": len(rows),
        "activations": len(activations) - dropped["trade_in"],
        "commission_income": round(sum(a["commission"] for a in activations), 2),
        "device_rebate": round(sum(a["device_rebate"] for a in activations), 2),
        "device_cost": round(sum(a["device_cost"] for a in activations), 2),
        "device_gp": round(sum(a["device_gp"] for a in activations), 2),
        "gross_profit_total": round(sum(a["device_gp"] for a in activations)
                                    + sum(a["commission"] for a in activations), 2),
        "distinct_customers": len({(a["phone"] or a["customer_name"]) for a in activations if (a["phone"] or a["customer_name"])}),
        "distinct_imeis": len({a["imei"] for a in activations if a["imei"]}),
        "dropped": dropped,
    }
    by_store_period = sorted(summ.values(), key=lambda s: (s["store_name"] or "", s["period"] or ""))
    # counts per Product Name family (for the confirm screen)
    families: dict[str, dict] = {}
    for a in activations:
        f = families.setdefault(a["product_name"] or "(blank)", {"product_name": a["product_name"], "kind": a["kind"], "count": 0, "amount": 0.0})
        f["count"] += 1
        f["amount"] = round(f["amount"] + a["unit_rebate"], 2)

    return {
        "activations": activations,
        "summary_by_store_period": by_store_period,
        "families": sorted(families.values(), key=lambda x: -x["amount"]),
        "totals": totals,
    }


# ── I/O: read the workbook into canonical rows ────────────────────────────────────────────────────
def read_xlsx(raw: bytes, sheet: str | None = None) -> list[dict]:
    """Read the first (or named) worksheet into a list of dicts keyed by the CANONICAL column name
    (see _COLS). Unknown columns are ignored. Returns [] on an unreadable file."""
    try:
        import io
        import openpyxl
    except Exception:
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception:
        return []
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    colmap = {}
    for i, h in enumerate(header or []):
        key = _COLS.get(str(h or "").strip().lower())
        if key:
            colmap[i] = key
    out = []
    for row in it:
        d = {}
        for i, key in colmap.items():
            if i < len(row):
                d[key] = row[i]
        if any(v not in (None, "") for v in d.values()):
            out.append(d)
    return out


def is_xlsx(raw: bytes) -> bool:
    # xlsx is a zip; magic bytes 'PK\x03\x04'
    return bool(raw) and raw[:4] == b"PK\x03\x04"


# ── Write path (customers + activations + P&L ledger) ─────────────────────────────────────────────
def _page_all(q, page=1000, cap=200000):
    out, start = [], 0
    while start < cap:
        rows = (q.range(start, start + page - 1).execute().data) or []
        out.extend(rows)
        if len(rows) < page:
            break
        start += page
    return out


def _zip5(z):
    d = re.sub(r"\D", "", str(z or ""))
    return d[:5] or None


def import_report(client, org_id: str, res: dict, *, store_code: str | None = None,
                  uploaded_by: str | None = None, source: str = "vendor_rebate_report") -> dict:
    """Idempotently write the parsed report:
      • CUSTOMERS  — insert new pos.customers deduped by phone (existing phones are left untouched).
      • ACTIVATIONS— insert pos.activations deduped by (cell_number, activation_date).
      • P&L LEDGER — UPSERT one commcalc.activation_rebate_ledger row per (store, period) with the
        commission + device-rebate totals (feeds coa's carrier_comm / device_rebate). Re-running the
        same file is a no-op; re-uploading a corrected period replaces that period's aggregate.
    Returns counts. Safe to re-run."""
    acts = res.get("activations") or []
    pos = client.schema("pos")

    # 1) customers — dedupe by phone_primary
    existing_phones = {_digits(r.get("phone_primary")) for r in
                       _page_all(pos.table("customers").select("phone_primary").eq("org_id", org_id))
                       if _digits(r.get("phone_primary"))}
    seen, new_customers = set(existing_phones), []
    for a in acts:
        ph = a.get("phone")
        if not ph or ph in seen:
            continue
        seen.add(ph)
        new_customers.append({
            "org_id": org_id, "first_name": a.get("first_name"), "last_name": a.get("last_name"),
            "phone_primary": ph, "zip": _zip5(a.get("zip")), "account_type": "individual", "is_active": True,
        })
    customers_created = 0
    for i in range(0, len(new_customers), 500):
        chunk = new_customers[i:i + 500]
        r = pos.table("customers").insert(chunk).execute()
        customers_created += len(r.data or [])

    # map phone -> customer_id (for linking activations), fetch after insert
    phone_to_id = {}
    for r in _page_all(pos.table("customers").select("id,phone_primary").eq("org_id", org_id)):
        p = _digits(r.get("phone_primary"))
        if p and r.get("id"):
            phone_to_id[p] = r["id"]

    # 2) activations — dedupe by (cell_number, activation_date)
    existing_keys = {(_digits(r.get("cell_number")), str(r.get("activation_date") or "")[:10])
                     for r in _page_all(pos.table("activations").select("cell_number,activation_date").eq("org_id", org_id))}
    new_acts = []
    for a in acts:
        if a.get("kind") == KIND_TRADE_IN:
            continue
        cell = a.get("cell_number")
        key = (_digits(cell), a.get("sold_on") or "")
        if key in existing_keys or not (cell or a.get("imei")):
            continue
        existing_keys.add(key)
        new_acts.append({
            "org_id": org_id, "store_code": store_code,
            "customer_id": phone_to_id.get(a.get("phone")),
            "carrier": a.get("carrier"), "activation_date": a.get("sold_on"),
            "plan_description": a.get("rate_plan"),
            "cell_number": cell, "mobile_phone": cell,
            "phone_serial": a.get("imei"), "phone_model": a.get("device_name"),
            "account_number": a.get("account_number"),
            "memo": " · ".join([x for x in (a.get("invoice_no"), a.get("product_name")) if x]) or None,
            "description": a.get("device_name"), "status": "active",
        })
    activations_created = 0
    for i in range(0, len(new_acts), 500):
        chunk = new_acts[i:i + 500]
        r = pos.table("activations").insert(chunk).execute()
        activations_created += len(r.data or [])

    # 3) P&L ledger — one aggregated row per (store, period); UPSERT so a re-upload replaces cleanly
    ledger_rows = []
    for s in (res.get("summary_by_store_period") or []):
        if not s.get("period"):
            continue
        ledger_rows.append({
            "org_id": org_id, "business_address": s.get("store_name"), "period": s["period"],
            "source": source, "commission_amount": s.get("commission_income") or 0,
            "device_rebate_amount": s.get("device_rebate") or 0, "device_cost": s.get("device_cost") or 0,
            "activations": s.get("activations") or 0,
        })
    ledger_upserted = 0
    if ledger_rows:
        r = (client.schema("commcalc").table("activation_rebate_ledger")
             .upsert(ledger_rows, on_conflict="org_id,business_address,period,source").execute())
        ledger_upserted = len(r.data or [])

    return {
        "customers_created": customers_created,
        "activations_created": activations_created,
        "ledger_periods": ledger_upserted,
        "totals": res.get("totals"),
    }
