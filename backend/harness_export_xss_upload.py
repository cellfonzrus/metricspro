"""ADVERSARIAL PROOF HARNESS — export formula injection (H7) · stored-XSS hrefs (H6) · upload cap (H5).

2026-08-05 app-wide security register, fix-order item ⑤ + the two shared-frontend findings.
Package: `agent/platform-core/export-xss-upload-hardening`.

No database, no network, no browser. Every section runs the REAL shipped code; where an attack is
claimed, the harness FIRST reproduces it against the base-`6aadb14` source pulled straight out of git
(a live negative control — if the vulnerability were imaginary the control fails and the harness stops).

SECTIONS
  A. H7 ATTACK, BACKEND   — openpyxl types a leading "=" as a real FORMULA. Reproduced on base
                            (`<f>` element in the saved xlsx), then proven dead after the fix.
  B. H7 ANTI-REGRESSION   — THE TRAP. Money, dates, phone numbers with a leading "+", and strings
                            legitimately starting with "-" survive byte-for-byte and keep their Excel
                            type. Proven by a CELL-GRID DIFF of base vs fixed `build_xlsx` over a
                            realistic money payload: identical value, type, number-format, alignment.
  C. H7 PDF               — the same renderer's PDF leg: reportlab Paragraph markup escaping. Base
                            CRASHES the entire PDF send when tenant text in the title / subtitle /
                            sheet name / column header contains an unbalanced markup tag; the data
                            cells were escaped, those four were not. Fixed here, with the crash
                            reproduced on base first.
  D. H6 hrefs             — `safe_href` allow-list, and the four TENANT-WRITABLE href fields this
                            module owns, each exercised through its real cleaner/endpoint.
  E. H5 upload cap        — the real ASGI middleware: declared Content-Length, lying Content-Length,
                            chunked bodies, every legitimate file size this app is known to ingest,
                            the break-glass, and GET/HEAD/OPTIONS untouched.
  F. H5 read_only         — REGISTER CORRECTION, evidenced: there is no direct `openpyxl.load_workbook`
                            in app code at all; every workbook read goes through pandas, which already
                            passes `read_only=True`.
  G. WIRING + BLAST RADIUS— middleware order, route count vs base, no money module touched, no SQL,
                            no SEED_VERSION bump, no new dependency.

Run:  cd backend && python3 harness_export_xss_upload.py
"""
import asyncio
import importlib.util
import io
import json
import os
import re
import subprocess
import sys
import zipfile

# Anchored to THIS FILE's directory, not the shell's cwd, so the harness runs identically from
# `backend/` and from the repo root (commit 564c171f).
_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _HERE)

PASS, FAIL = 0, 0
BASE_REF = "6aadb14"
REPO = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))


def ok(name, cond, extra=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ok   {name}")
    else:
        FAIL += 1
        print(f"  FAIL {name} {extra}")


def fatal(msg):
    print(f"\nFATAL: {msg}")
    sys.exit(2)


# ── reportlab: SKIP (loudly) vs PRODUCT DEFECT ───────────────────────────────────────────────────
# Section C below died on `ModuleNotFoundError: No module named 'reportlab'`. Before making it
# tolerate that, the question that actually matters was answered: does the PRODUCT need reportlab?
#
#   It needs it    — app/modules/notify/render.py::build_pdf (the module under test here) imports it
#                    lazily inside the render call, as do three other shipped modules. Live paths.
#   It declares it — backend/requirements.txt line 13, `reportlab>=4.2.0`, uncommented.
#   It installs it — backend/Dockerfile: `RUN pip install --no-cache-dir -r requirements.txt`.
#
# So the deployed image HAS reportlab; no production defect. What is missing is THIS container's
# copy (it is short 7 declared deps). A skip is the honest answer — but only for that case:
#   installed              -> RUN section C for real.
#   missing but DECLARED   -> SKIP, counted and printed in the summary. Never a silent pass.
#   missing and UNDECLARED -> FAIL loudly: shipped code importing an undeclared package is a
#                             production defect and this harness will say so.
#
# WHY A PLAIN try/except WOULD HAVE BEEN WORSE THAN THE CRASH. C1 is a NEGATIVE control: it asserts
# the pre-fix renderer CRASHES on tenant markup, by catching any Exception. `ModuleNotFoundError` is
# an Exception, so with reportlab absent C1 was recorded as a PASS — the control "fired" without the
# code under test ever being reached. Every skip below therefore replaces BOTH halves of the pair.
SKIPPED = []


def _pdf_backend():
    """('run' | 'skip' | 'defect', message) for the reportlab PDF backend."""
    import importlib.util
    installed = importlib.util.find_spec("reportlab") is not None
    declared = False
    try:
        for line in open(os.path.join(_HERE, "requirements.txt"), encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and line.split("[")[0].split("=")[0] \
                    .split(">")[0].split("<")[0].strip().lower() == "reportlab":
                declared = True
                break
    except OSError:
        pass
    if installed:
        return "run", "reportlab present"
    if declared:
        return "skip", ("reportlab is DECLARED in backend/requirements.txt and installed by the "
                        "Dockerfile, but is absent from THIS container — environment gap, not a "
                        "product defect. Install it to run section C: pip install reportlab")
    return "defect", ("PRODUCT DEFECT: shipped code imports reportlab but backend/requirements.txt "
                      "does not declare it — the deployed image would 500 on every PDF export")


PDF_MODE, PDF_WHY = _pdf_backend()


def skip(name, why):
    SKIPPED.append(name)
    print(f"  SKIP {name}\n       {why}")


def section(t):
    print(f"\n── {t} " + "─" * max(0, 88 - len(t)))


VECTORS = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                      "harness_vectors_export_xss.json"), encoding="utf-8"))
RISKY = VECTORS["formula_risky"]["vectors"]
SAFE_CELLS = VECTORS["formula_safe"]["vectors"]
BAD_HREF = VECTORS["href_unsafe"]["vectors"]
GOOD_HREF = VECTORS["href_safe"]["vectors"]


def git_module(rel_path, mod_name):
    """Import a file exactly as it exists at BASE_REF — the negative control's source of truth."""
    try:
        src = subprocess.run(["git", "-C", REPO, "show", f"{BASE_REF}:{rel_path}"],
                             capture_output=True, check=True).stdout.decode()
    except Exception as exc:                                   # pragma: no cover
        fatal(f"could not read {rel_path} at {BASE_REF}: {exc}")
    path = os.path.join("/tmp", f"_base_{mod_name}.py")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(src)
    spec = importlib.util.spec_from_file_location(mod_name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ══ helpers ══════════════════════════════════════════════════════════════════════════════════════
def sheet_xml(xlsx_bytes):
    z = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    return "".join(z.read(n).decode("utf-8", "replace")
                   for n in z.namelist() if n.startswith("xl/worksheets/"))


def grid(xlsx_bytes):
    """Everything a human SEES, plus the cell's Excel type — the report's displayed values."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None and c.data_type == "n":
                    continue
                out.append((ws.title, c.coordinate, repr(c.value), c.data_type,
                            c.number_format,
                            (c.alignment.horizontal if c.alignment else None)))
    return out


def payload_of(values, money_values=None):
    cols = [{"header": "Store", "key": "store"}, {"header": "Note", "key": "note"}]
    if money_values is not None:
        cols.append({"header": "Commission", "key": "amt", "money": True})
    rows = []
    for i, v in enumerate(values):
        r = {"store": "Jamaica Ave", "note": v}
        if money_values is not None:
            r["amt"] = money_values[i % len(money_values)]
        rows.append(r)
    return {"title": "Rep Commission", "subtitle": "July 2026", "filename": "rep",
            "sheets": [{"name": "Detail", "columns": cols, "rows": rows}]}


# ══ A. H7 ATTACK — BACKEND xlsx ══════════════════════════════════════════════════════════════════
section("A. H7 — CSV/Excel formula injection, BACKEND export path (notify/render.py)")

BASE_RENDER = git_module("backend/app/modules/notify/render.py", "base_render")
from app.modules.notify import render as R                                        # noqa: E402

# A0 — the primitive itself is a live formula. Not an assumption; openpyxl's own typing.
from openpyxl import Workbook                                                     # noqa: E402
_wb = Workbook()
ok("A0  openpyxl types a leading '=' as a FORMULA (the root cause)",
   _wb.active.cell(row=1, column=1, value="=cmd|'/C calc'!A0").data_type == "f")
ok("A0b openpyxl types '-1234.56' (a string) as TEXT, not a formula",
   _wb.active.cell(row=2, column=1, value="-1234.56").data_type == "s")

# A1 — NEGATIVE CONTROL: base build_xlsx emits <f> for the payloads that begin with '='.
eq_vectors = [v for v in RISKY if v.startswith("=")]
base_xml = sheet_xml(BASE_RENDER.build_xlsx(payload_of(eq_vectors)))
ok(f"A1  NEGATIVE CONTROL: base @{BASE_REF} writes {len(eq_vectors)} live <f> formula cells",
   base_xml.count("<f>") == len(eq_vectors), f"found {base_xml.count('<f>')}")
ok("A1b NEGATIVE CONTROL: the DDE payload is present as a formula on base",
   "<f>cmd|" in base_xml.replace("<f>=", "<f>"))

# A2 — after the fix: zero formulas, for EVERY risky vector.
fixed = R.build_xlsx(payload_of(RISKY))
fixed_xml = sheet_xml(fixed)
ok(f"A2  FIXED: not one <f> element for any of the {len(RISKY)} attack vectors",
   "<f>" not in fixed_xml and "<f " not in fixed_xml)

# A3 — and the displayed characters did not change (that is the whole design).
from openpyxl import load_workbook                                                # noqa: E402
_ws = load_workbook(io.BytesIO(R.build_xlsx(payload_of(RISKY)))).worksheets[0]
_wsb = load_workbook(io.BytesIO(BASE_RENDER.build_xlsx(payload_of(RISKY)))).worksheets[0]
readback = [_ws.cell(row=ri, column=2).value for ri in range(2, len(RISKY) + 2)]
base_readback = [_wsb.cell(row=ri, column=2).value for ri in range(2, len(RISKY) + 2)]
ok("A3  FIXED: every attack payload reads back EXACTLY as it does on base (nothing rewritten)",
   readback == base_readback,
   f"{[(a, b) for a, b in zip(readback, base_readback) if a != b][:2]}")
# ...and that base readback is the input itself, except for one pre-existing XML rule: a leading
# CARRIAGE RETURN is normalised to LF by the XML spec on parse. That is openpyxl/XML, on base too.
diffs = [(a, b) for a, b in zip(base_readback, RISKY) if a != b]
ok("A3b the only value openpyxl cannot round-trip is a leading CR (XML line-ending "
   "normalisation) — pre-existing on base, not introduced here",
   diffs == [("\n=1+1", "\r=1+1")], diffs)
ok("A3c FIXED: every neutralised cell is a TEXT cell",
   all(_ws.cell(row=ri, column=2).data_type == "s" for ri in range(2, len(RISKY) + 2)))
ok("A3d FIXED: every neutralised cell carries Excel's quotePrefix marker",
   all(_ws.cell(row=ri, column=2).quotePrefix for ri in range(2, len(RISKY) + 2)))

# A4 — a malicious COLUMN HEADER (headers are configurable on custom reports) is covered too.
hdr_payload = {"title": "t", "filename": "f", "sheets": [{"name": "S", "rows": [{"k": 1}],
               "columns": [{"header": "=cmd|'/C calc'!A0", "key": "k"}]}]}
ok("A4  NEGATIVE CONTROL: base makes a malicious column HEADER a formula",
   "<f>" in sheet_xml(BASE_RENDER.build_xlsx(hdr_payload)))
ok("A4b FIXED: a malicious column header is inert",
   "<f>" not in sheet_xml(R.build_xlsx(hdr_payload)))

# A5 — the classifier, over the whole corpus.
ok(f"A5  _is_formula_risky flags all {len(RISKY)} attack vectors",
   all(R._is_formula_risky(v) for v in RISKY),
   [v for v in RISKY if not R._is_formula_risky(v)][:3])
ok(f"A5b _is_formula_risky clears all {len(SAFE_CELLS)} legitimate vectors",
   not any(R._is_formula_risky(v) for v in SAFE_CELLS),
   [v for v in SAFE_CELLS if R._is_formula_risky(v)][:3])
ok("A5c non-strings can never be risky (int/float/bool/None/date)",
   not any(R._is_formula_risky(v) for v in
           [0, -1234.56, 1, True, False, None, __import__("datetime").date(2026, 8, 6)]))


# ══ B. H7 ANTI-REGRESSION — the money/date/phone trap ════════════════════════════════════════════
section("B. H7 ANTI-REGRESSION — no report's displayed values move")

MONEY = [-1234.56, 0, 27043.19, -0.01, 124043.34]
legit = payload_of(SAFE_CELLS, MONEY)
gb, gf = grid(BASE_RENDER.build_xlsx(legit)), grid(R.build_xlsx(legit))
ok("B1  CELL-GRID IDENTICAL to base for a realistic money report (value+type+format+align)",
   gb == gf, f"{[x for x in zip(gb, gf) if x[0] != x[1]][:2]}")

# B2 — negative currency stays a NUMBER with the money format (the regression the brief warns about).
wb_f = load_workbook(io.BytesIO(R.build_xlsx(legit)))
ws_f = wb_f.worksheets[0]
money_cells = [ws_f.cell(row=r, column=3) for r in range(2, 2 + len(SAFE_CELLS))]
ok("B2  every money cell is still a NUMBER (data_type 'n'), never text",
   all(c.data_type == "n" for c in money_cells))
ok("B2b every money cell keeps the $#,##0.00 number format",
   all(c.number_format == R.MONEY_FMT for c in money_cells))
ok("B2c negative currency is stored as a negative NUMBER, unrounded",
   any(c.value == -1234.56 for c in money_cells) and any(c.value == 124043.34 for c in money_cells))
ok("B2d money cells are NOT quote-prefixed (nothing about them changed)",
   not any(c.quotePrefix for c in money_cells))

# B3 — the four named regression classes, individually, in a TEXT column.
CASES = {
    "negative currency string": "-1234.56",
    "negative currency w/ symbol+group": "-$1,234.56",
    "ISO date": "2026-08-06",
    "US date": "08/06/2026",
    "phone with leading +": "+1 (555) 123-4567",
    "string legitimately starting with '-'": "-Adjustment",
}
one = R.build_xlsx(payload_of(list(CASES.values())))
ws1 = load_workbook(io.BytesIO(one)).worksheets[0]
for i, (label, val) in enumerate(CASES.items(), start=2):
    cell = ws1.cell(row=i, column=2)
    ok(f"B3  {label}: exported characters unchanged", cell.value == val, repr(cell.value))
# ...and their Excel typing is right: the numeric/date ones are untouched text as before the fix,
# the genuinely-risky ones are text too — in NO case does a value gain or lose a character.
base1 = load_workbook(io.BytesIO(BASE_RENDER.build_xlsx(payload_of(list(CASES.values()))))).worksheets[0]
ok("B3b every one of those six cells reads IDENTICALLY to base",
   [ws1.cell(row=i, column=2).value for i in range(2, 8)]
   == [base1.cell(row=i, column=2).value for i in range(2, 8)])
ok("B3c the numeric-looking ones are left completely alone (no quotePrefix added)",
   not ws1.cell(row=2, column=2).quotePrefix and not ws1.cell(row=3, column=2).quotePrefix)
ok("B3d the genuinely risky ones ARE marked (phone / '-Adjustment')",
   ws1.cell(row=6, column=2).quotePrefix and ws1.cell(row=7, column=2).quotePrefix)

# B4 — a report with no risky data at all is grid-identical to base (the common case: every report).
plain = payload_of(["Jamaica Ave", "Queens Blvd", "Store 12"], MONEY)
ok("B4  an ordinary report is grid-identical to base",
   grid(BASE_RENDER.build_xlsx(plain)) == grid(R.build_xlsx(plain)))

# B5 — money is a float BEFORE it can ever reach the guard: no money value can be neutralised.
src_render = open(os.path.join(_HERE, "app/modules/notify/render.py"), encoding="utf-8").read()
xlsx_body = src_render.split("def build_xlsx")[1].split("def build_pdf")[0]
money_branch = xlsx_body.split('if col.get("money"):')[1].split("else:")[0]
ok("B5  the build_xlsx MONEY branch still writes a float + number format",
   "float(_raw(col, row) or 0)" in money_branch and "MONEY_FMT" in money_branch)
ok("B5b the money branch contains no call to the neutraliser — money cannot be touched",
   "_is_formula_risky" not in money_branch and "_write_text_cell" not in money_branch)


# ══ C. H7 — the PDF leg of the same renderer ═════════════════════════════════════════════════════
section("C. PDF leg — reportlab Paragraph markup (pre-existing crash, fixed in passing)")

# Data cells were already escaped on base. The TITLE, SUBTITLE, SHEET NAME and COLUMN HEADERS were
# not — and all four carry tenant text. reportlab parses Paragraph mini-markup, so any tenant string
# containing a KNOWN markup tag (<b> <i> <u> <br> <font> <super> ...) that is not balanced makes
# doc.build() RAISE and the entire PDF send fail for that tenant.
MARKUP = "Rate <b of 2"          # e.g. a note/label an admin typed; base cannot render it
PDF_CASES = [
    ("title", {"title": MARKUP, "filename": "f",
               "sheets": [{"name": "S", "columns": [{"header": "H", "key": "k"}], "rows": [{"k": 1}]}]}),
    ("subtitle", {"title": "t", "subtitle": MARKUP, "filename": "f",
                  "sheets": [{"name": "S", "columns": [{"header": "H", "key": "k"}], "rows": [{"k": 1}]}]}),
    ("column header", {"title": "t", "filename": "f",
                       "sheets": [{"name": "S", "columns": [{"header": MARKUP, "key": "k"}],
                                   "rows": [{"k": 1}]}]}),
    ("sheet name", {"title": "t", "filename": "f", "sheets": [
        {"name": MARKUP, "columns": [{"header": "H", "key": "k"}], "rows": [{"k": 1}]},
        {"name": "S2", "columns": [{"header": "H", "key": "k"}], "rows": [{"k": 1}]}]}),
]

if PDF_MODE == "defect":
    ok("C0  reportlab is DECLARED wherever shipped code imports it", False, PDF_WHY)
elif PDF_MODE == "skip":
    for where, _pay in PDF_CASES:
        skip(f"C1  NEGATIVE CONTROL: base CRASHES the whole PDF on tenant markup in the {where}",
             PDF_WHY)
        skip(f"C1b FIXED: the same report renders a real PDF ({where})", PDF_WHY)
    skip("C2  an ordinary PDF still renders", PDF_WHY)
    skip("C3  a formula payload cannot execute from a PDF (text only) and still renders", PDF_WHY)
else:
    for where, pay in PDF_CASES:
        base_ok = True
        try:
            BASE_RENDER.build_pdf(pay)
        except ModuleNotFoundError:          # never let a missing dep masquerade as "base crashes"
            fatal("reportlab vanished mid-run — the C1 negative control cannot be trusted")
        except Exception:
            base_ok = False
        ok(f"C1  NEGATIVE CONTROL: base CRASHES the whole PDF on tenant markup in the {where}",
           not base_ok)
        try:
            fixed_ok = R.build_pdf(pay).startswith(b"%PDF")
        except Exception as exc:                                                  # pragma: no cover
            fixed_ok = False
            print("        ", exc)
        ok(f"C1b FIXED: the same report renders a real PDF ({where})", fixed_ok)
    ok("C2  an ordinary PDF still renders",
       R.build_pdf(payload_of(["Jamaica"], MONEY)).startswith(b"%PDF"))
    ok("C3  a formula payload cannot execute from a PDF (text only) and still renders",
       R.build_pdf(payload_of(RISKY, MONEY)).startswith(b"%PDF"))

# Reportlab-INDEPENDENT half of section C, so a missing PDF backend never takes the whole H7 proof
# down with it: the FIX itself is pure string work and is readable from source either way.
_RENDER_SRC = open(os.path.join(_HERE, "app/modules/notify/render.py"), encoding="utf-8").read()
assert "def _para(" in _RENDER_SRC, "render.py::_para is gone — this anchor moved, fix the harness"
_para_body = _RENDER_SRC.split("def _para(")[1].split("\ndef ")[0]
ok("C4  (backend-independent) the Paragraph escape helper covers all three metacharacters",
   all(m in _para_body for m in ('"&", "&amp;"', '"<", "&lt;"', '">", "&gt;"')), _para_body[-160:])
_pdf_body = _RENDER_SRC.split("def build_pdf")[1]
ok("C5  (backend-independent) build_pdf applies it to the FOUR tenant-text sites the fix names "
   "(title, subtitle, sheet name, column header), not only to data cells",
   all(f"_para({site}" in _pdf_body.replace(" ", "") or site in _pdf_body
       for site in ('payload.get("title")', 'payload["subtitle"]',
                    "sheet.get('name')", 'c.get("header")'))
   and _pdf_body.count("_para(") >= 4, _pdf_body.count("_para("))


# ══ D. H6 — stored XSS via javascript: hrefs ═════════════════════════════════════════════════════
section("D. H6 — tenant-writable hrefs (allow-list at the WRITE side)")

from app.modules.core.safe_href import safe_href, is_safe_href                     # noqa: E402

ok(f"D1  is_safe_href rejects all {len(BAD_HREF)} XSS/off-site vectors",
   not any(is_safe_href(v) for v in BAD_HREF), [v for v in BAD_HREF if is_safe_href(v)][:3])
ok(f"D1b is_safe_href accepts all {len(GOOD_HREF)} real link shapes",
   all(is_safe_href(v) for v in GOOD_HREF), [v for v in GOOD_HREF if not is_safe_href(v)][:3])
ok("D1c safe_href is NON-REWRITING (a legitimate href comes back byte-identical)",
   all(safe_href(v) == v for v in GOOD_HREF))
ok("D1d safe_href drops an unsafe value to None (link simply does not render)",
   all(safe_href(v) is None for v in BAD_HREF))
ok("D1e empty / None / whitespace are not links", not any(is_safe_href(v) for v in ["", None, "   "]))

# D2 — the REAL cleaners for the two tour fields (the auto-firing `?tour=` path).
from app.modules.core import training as TR                                        # noqa: E402
ok("D2  clean_step strips a javascript: page_href",
   TR.clean_step({"title": "t", "body": "b", "page_href": "javascript:alert(1)"}, 0)["page_href"] is None)
ok("D2b clean_tour strips a javascript: start_href",
   TR.clean_tour({"title": "t", "start_href": "JaVaScRiPt:alert(1)"})["start_href"] is None)
ok("D2c clean_step keeps every legitimate page_href unchanged",
   all(TR.clean_step({"title": "t", "body": "b", "page_href": h}, 0)["page_href"] == h
       for h in GOOD_HREF))

# D2d — the SHIPPED tour pack must survive the new gate untouched (real product data, not a fixture).
seed = json.load(open(os.path.join(_HERE, "app/data/training_tours_seed.json"), encoding="utf-8"))
seed_hrefs = [t.get("start_href") for t in seed["tours"] if t.get("start_href")]
seed_hrefs += [s.get("page_href") for t in seed["tours"] for s in t.get("steps", []) if s.get("page_href")]
ok(f"D2d all {len(seed_hrefs)} hrefs in the SHIPPED tour pack survive unchanged",
   all(safe_href(h) == h for h in seed_hrefs), [h for h in seed_hrefs if safe_href(h) != h][:3])

# D3 — What's New deep_link.
from app.modules.core import whats_new as WN                                       # noqa: E402
ok("D3  clean_entry strips a javascript: deep_link",
   WN.clean_entry({"title": "t", "deep_link": "javascript:alert(1)"})["deep_link"] is None)
ok("D3b clean_entry keeps a real deep_link",
   WN.clean_entry({"title": "t", "deep_link": "/closing/envelope-payout"})["deep_link"]
   == "/closing/envelope-payout")

# D4 — the Import Health feed registry (admin-editable deep link) via its REAL endpoint.
from app.modules.core import import_health as IH                                   # noqa: E402


class FakeTable:
    def __init__(self, sink):
        self.sink = sink

    def insert(self, row):
        self.sink.append(row)
        return self

    def execute(self):
        return type("R", (), {"data": [self.sink[-1]]})()


class FakeClient:
    def __init__(self):
        self.rows = []

    def schema(self, _):
        return self

    def table(self, _):
        return FakeTable(self.rows)


fc = FakeClient()
_orig_gate = IH._gate
IH._gate = lambda *a, **k: (fc, {"role": "admin"}, "00000000-0000-0000-0000-0000000000ff")
try:
    # FastAPI hands a POST handler its DECLARED, validated pydantic model — never a bare dict.
    # `create_import_feed` reads `body.feed_key`, so a dict raised AttributeError and the handler
    # blew up BEFORE reaching the deep_link check these two assertions exist to prove. Same drift
    # and same repair as commit 564c171f (harness_ssrf_import_gate).
    IH.create_import_feed(IH.CreateImportFeedIn(label="x", deep_link="javascript:alert(1)"))
    ok("D4  POST /core/import-feeds falls back to the safe default for a javascript: deep_link",
       fc.rows[-1]["deep_link"] == "/commcalc/upload", fc.rows[-1]["deep_link"])
    IH.create_import_feed(IH.CreateImportFeedIn(label="x", deep_link="/commcalc/ftp-imports"))
    ok("D4b a real deep_link is stored unchanged", fc.rows[-1]["deep_link"] == "/commcalc/ftp-imports")
finally:
    IH._gate = _orig_gate

# D5 — portal_reports.href (rendered on the employee portal) via its REAL endpoint.
from fastapi import HTTPException                                                  # noqa: E402
from app.modules.core import router as CR                                          # noqa: E402
_orig_req = CR._require_setting
CR._require_setting = lambda *a, **k: None
try:
    # Was passing a bare dict; FastAPI hands this handler its declared `SetPortalReportIn`, so every
    # probe died on AttributeError — and the bare `except Exception: pass` below SWALLOWED it, so
    # `rejected` stayed 0 and this security assertion had silently stopped guarding the href gate
    # entirely. The blanket except is now narrowed: anything that is not an HTTPException is a
    # harness fault and is reported, never counted and never quietly ignored.
    rejected, faults = 0, []
    for v in BAD_HREF:
        try:
            CR.set_portal_report(CR.SetPortalReportIn(href=v), org_id="x")
        except HTTPException as e:
            rejected += 1 if e.status_code == 400 else 0
        except Exception as exc:                                                  # pragma: no cover
            faults.append(f"{v!r} -> {type(exc).__name__}: {exc}")
    ok("D5a the probes actually reached the endpoint (no harness-side error was swallowed)",
       not faults, "; ".join(faults[:3]))
    ok(f"D5  PUT /core/portal-reports rejects all {len(BAD_HREF)} unsafe hrefs with 400",
       rejected == len(BAD_HREF), f"{rejected}/{len(BAD_HREF)}")
finally:
    CR._require_setting = _orig_req

# D6 — the render-side sinks are wired (source-level; the TS behaviour is proven in the .mjs twin).
FE = os.path.join(REPO, "frontend", "src")


def read(p):
    return open(os.path.join(FE, p), encoding="utf-8").read()


tr = read("../src/components/TourRunner.tsx") if False else read("components/TourRunner.tsx")
ok("D6  TourRunner refuses to auto-navigate to an unsafe page_href",
   "isSafeHref(step.page_href)" in tr and tr.index("isSafeHref(step.page_href)") < tr.index("router.push(step.page_href)"))
tours = read("lib/tours.ts")
ok("D6b tours.ts scrubs hrefs the moment they arrive (fetchTour + fetchTours)",
   "sanitizeTourHrefs" in tours and "safeHref(t.start_href)" in tours)
root_layout = read("app/layout.tsx")
ok("D6c the app-wide click net is mounted in the ROOT layout (covers /portal + /onboard/[token])",
   "UnsafeLinkGuard" in root_layout)
for f, needle in [("components/AdminAttention.tsx", "safeHref(n.deep_link"),
                  ("components/AdminAttention.tsx", "safeHref(it.deep_link"),
                  ("components/PortalReports.tsx", "safeHref(r.href"),
                  ("app/(platform)/admin/import-health/page.tsx", "safeHref(i.deep_link"),
                  ("app/(platform)/admin/import-health/page.tsx", "safeHref(f.deep_link"),
                  ("app/(platform)/remediation/page.tsx", "safeHref(result.approval_url"),
                  ("app/(platform)/training/page.tsx", "safeHref(t.start_href")]:
    ok(f"D6  sink sanitised: {f} :: {needle})", needle in read(f))


# ══ E. H5 — request-body size cap ════════════════════════════════════════════════════════════════
section("E. H5 — request-body size cap (backend/app/core/body_limit.py)")

from app.core import body_limit as BL                                              # noqa: E402

BL._log_oversize = lambda *a, **k: None      # keep the harness DB-free


def call(method="POST", chunks=(b"",), content_length=None, path="/api/v1/commcalc/upload"):
    """Drive the REAL ASGI middleware. Returns (status, inner_app_saw_request, body_bytes_delivered)."""
    seen = {"called": False, "bytes": 0}

    async def inner(scope, receive, send):
        seen["called"] = True
        while True:
            m = await receive()
            if m["type"] != "http.request":
                break
            seen["bytes"] += len(m.get("body") or b"")
            if not m.get("more_body"):
                break
        body = b'{"ok":true}'
        await send({"type": "http.response.start", "status": 200,
                    "headers": [(b"content-type", b"application/json")]})
        await send({"type": "http.response.body", "body": body})

    headers = []
    if content_length is not None:
        headers.append((b"content-length", str(content_length).encode()))
    scope = {"type": "http", "method": method, "path": path, "headers": headers,
             "query_string": b""}
    it = iter(list(chunks))

    async def receive():
        try:
            b = next(it)
        except StopIteration:
            return {"type": "http.disconnect"}
        more = True
        try:
            nxt = next(it)
            it2 = iter([nxt] + list(it))
        except StopIteration:
            more = False
            it2 = iter([])
        # rebuild the iterator so the peek is not consumed
        for _ in ():
            pass
        return {"type": "http.request", "body": b, "more_body": more, "_it": it2}

    # simpler, deterministic receive: emit each chunk, last one with more_body=False
    chunk_list = list(chunks)
    idx = {"i": 0}

    async def receive2():
        i = idx["i"]
        if i >= len(chunk_list):
            return {"type": "http.disconnect"}
        idx["i"] += 1
        return {"type": "http.request", "body": chunk_list[i],
                "more_body": idx["i"] < len(chunk_list)}

    out = {"status": None}

    async def send(message):
        if message["type"] == "http.response.start":
            out["status"] = message["status"]

    asyncio.run(BL.BodySizeLimitMiddleware(inner)(scope, receive2, send))
    return out["status"], seen["called"], seen["bytes"]


MB = 1024 * 1024
ok("E0  default cap is 64 MB", BL.max_upload_bytes() == 64 * MB, BL.max_upload_bytes())

# E1 — the evidence: every upload size this app is KNOWN to handle passes untouched.
EVIDENCE = [
    ("full-month Sales Transaction Details workbook (documented, HANDOFF.md)", 7 * MB),
    ("asset_ledger full re-upload, 43,849 rows x 30 cols, high-entropy worst case", int(9.49 * MB)),
    ("hostile synthetic 50,000 x 78 sales workbook (larger than anything real)", int(28.32 * MB)),
    ("phone-camera evidence photo, high resolution", 15 * MB),
    ("a report emailed back through /notify/send-file, base64 (+33%)", 12 * MB),
]
for label, size in EVIDENCE:
    st, called, _ = call(content_length=size, chunks=(b"x" * 1024,))
    ok(f"E1  ALLOWED @ {size/MB:5.2f} MB — {label}", st == 200 and called)

# E2 — the attack.
st, called, _ = call(content_length=200 * MB, chunks=(b"x" * 1024,))
ok("E2  200 MB declared body → 413, and the app is NEVER entered", st == 413 and not called)
st, called, _ = call(content_length=2 * 1024 * MB, chunks=(b"x",))
ok("E2b 2 GB declared body → 413 before a single byte is read", st == 413 and not called)

# E3 — a LYING Content-Length (or none at all): the streaming counter still stops it.
big = [b"x" * (4 * MB)] * 20                                    # 80 MB actually sent
st, called, got = call(content_length=1024, chunks=big)
ok("E3  Content-Length lies (1 KB declared, 80 MB sent) → 413 mid-stream", st == 413)
ok("E3b the app never received more than the cap", got <= 64 * MB, got)
st, called, got = call(content_length=None, chunks=big)
ok("E3c no Content-Length at all (chunked) → 413 mid-stream", st == 413)
st, called, got = call(content_length=None, chunks=[b"x" * (1 * MB)] * 7)
ok("E3d chunked 7 MB with no Content-Length passes through intact", st == 200 and got == 7 * MB)

# E4 — bodiless methods are untouched.
for m in ("GET", "HEAD", "OPTIONS"):
    st, called, _ = call(method=m, content_length=500 * MB, chunks=(b"",))
    ok(f"E4  {m} is never policed (no body to police)", st == 200 and called)

# E5 — knobs.
os.environ["MAX_UPLOAD_MB"] = "0"
st, called, _ = call(content_length=500 * MB, chunks=(b"x" * 16,))
ok("E5  MAX_UPLOAD_MB=0 is a true break-glass — pure pass-through", st == 200 and called)
os.environ["MAX_UPLOAD_MB"] = "128"
ok("E5b MAX_UPLOAD_MB tunes the cap without a code change", BL.max_upload_bytes() == 128 * MB)
st, called, _ = call(content_length=100 * MB, chunks=(b"x" * 16,))
ok("E5c a raised cap admits a bigger file", st == 200 and called)
os.environ["MAX_UPLOAD_MB"] = "garbage"
ok("E5d a malformed MAX_UPLOAD_MB falls back to the 64 MB default, never crashes",
   BL.max_upload_bytes() == 64 * MB)
os.environ.pop("MAX_UPLOAD_MB", None)

# E7 — END-TO-END through a REAL FastAPI multipart upload endpoint, not just a stub app. This is
# the assertion that matters: the abort must unwind through FastAPI's form parser, the router and
# ExceptionMiddleware and still come out as a clean 413, never a masked 500.
from fastapi import FastAPI, File, UploadFile                                      # noqa: E402

_real = FastAPI()


@_real.post("/upload")
async def _upload(file: UploadFile = File(...)):
    return {"bytes": len(await file.read())}


def multipart(nbytes):
    b = "----mpX"
    head = (f"--{b}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"x.xlsx\"\r\n"
            "Content-Type: application/vnd.ms-excel\r\n\r\n").encode()
    return b, head + (b"A" * nbytes) + f"\r\n--{b}--\r\n".encode()


def post_real(nbytes, declare=True, chunk=1 * 1024 * 1024):
    boundary, body = multipart(nbytes)
    headers = [(b"content-type", f"multipart/form-data; boundary={boundary}".encode())]
    if declare:
        headers.append((b"content-length", str(len(body)).encode()))
    parts = [body[i:i + chunk] for i in range(0, len(body), chunk)] or [b""]
    idx = {"i": 0}
    out = {"status": None, "body": b""}

    async def receive():
        i = idx["i"]
        if i >= len(parts):
            return {"type": "http.disconnect"}
        idx["i"] += 1
        return {"type": "http.request", "body": parts[i], "more_body": idx["i"] < len(parts)}

    async def send(m):
        if m["type"] == "http.response.start":
            out["status"] = m["status"]
        elif m["type"] == "http.response.body":
            out["body"] += m.get("body") or b""

    scope = {"type": "http", "asgi": {"version": "3.0", "spec_version": "2.1"}, "http_version": "1.1",
             "method": "POST", "scheme": "http", "path": "/upload", "raw_path": b"/upload",
             "root_path": "", "query_string": b"", "headers": headers,
             "client": ("1.2.3.4", 1234), "server": ("t", 80)}
    asyncio.run(BL.BodySizeLimitMiddleware(_real)(scope, receive, send))
    return out["status"], out["body"]


st, body = post_real(2 * MB)
ok("E7  REAL FastAPI multipart upload, 2 MB → 200 and the file arrives intact",
   st == 200 and b'"bytes":2097152' in body.replace(b" ", b""), (st, body[:120]))
st, body = post_real(80 * MB)
ok("E7b REAL FastAPI multipart upload, 80 MB declared → clean 413 (never a masked 500)",
   st == 413 and b"request_too_large" in body, (st, body[:120]))
st, body = post_real(80 * MB, declare=False)
ok("E7c REAL FastAPI multipart upload, 80 MB CHUNKED (no Content-Length) → clean 413 mid-parse",
   st == 413 and b"request_too_large" in body, (st, body[:120]))
os.environ["MAX_UPLOAD_MB"] = "0"
st, body = post_real(2 * MB)
ok("E7d break-glass OFF still serves a real upload unchanged", st == 200)
os.environ.pop("MAX_UPLOAD_MB", None)

# E6 — the 413 body is actionable and leaks nothing.
msg = BL._too_large_body(64 * MB).decode()
ok("E6  the 413 names the limit and the knob", "64 MB" in msg and "MAX_UPLOAD_MB" in msg)
ok("E6b the 413 leaks no path, stack or internal detail",
   "Traceback" not in msg and "/app/" not in msg)


# ══ F. H5 — the read_only=True half of the finding ═══════════════════════════════════════════════
section("F. H5 — `read_only=True`: REGISTER CORRECTION, with evidence")

app_py = subprocess.run(["grep", "-rn", "--include=*.py", "load_workbook",
                         os.path.join(REPO, "backend", "app")],
                        capture_output=True).stdout.decode()
_lw_sites = [ln for ln in app_py.splitlines() if "load_workbook(" in ln and "def " not in ln]
# F1 asserted "there are NO direct load_workbook calls in backend/app — nothing to add read_only to".
# That was true when written; two have since been added (commcalc/carrier_recon.py,
# pos/vendor_rebate_report.py) and the absolute went red while the HARDENING IT PROTECTS is
# perfectly healthy — both new sites pass read_only=True. Same stale-literal class as the provider
# list re-expressed in commit 564c171f: an absolute that says "none exist" cannot survive the first
# legitimate one, and a harness that cries wolf gets ignored.
#
# The invariant that actually matters (finding H5) is that a workbook is never loaded with the whole
# file materialised in memory — i.e. EVERY direct call passes read_only=True. Asserted as that
# property, adding a third safe call site can no longer break it, and adding an UNSAFE one now
# fails, which the old absolute could not distinguish.
ok("F1  every direct openpyxl.load_workbook in backend/app passes read_only=True (H5 hardening)",
   all("read_only=True" in ln for ln in _lw_sites),
   [ln for ln in _lw_sites if "read_only=True" not in ln])
ok("F1b …and data_only=True, so no formula is ever evaluated on load",
   all("data_only=True" in ln for ln in _lw_sites),
   [ln for ln in _lw_sites if "data_only=True" not in ln])
ok("F1c the F1 scan is non-vacuous (it really found the call sites it is judging)",
   len(_lw_sites) >= 2, _lw_sites)

import pandas as _pd                                                               # noqa: E402
import inspect as _inspect                                                         # noqa: E402
from pandas.io.excel._openpyxl import OpenpyxlReader                                # noqa: E402
src_lw = _inspect.getsource(OpenpyxlReader.load_workbook)
ok(f"F2  pandas {_pd.__version__} ALREADY loads every workbook with read_only=True",
   '"read_only": True' in src_lw)
ok("F2b ...and with data_only=True + keep_links=False (no formula eval, no external links)",
   '"data_only": True' in src_lw and '"keep_links": False' in src_lw)

writes_only = open(os.path.join(_HERE, "app/modules/notify/render.py"), encoding="utf-8").read()
ok("F3  the ONE direct openpyxl use in this codebase is a WRITE (Workbook()), where read_only "
   "does not apply", "from openpyxl import Workbook" in writes_only and
   "load_workbook" not in writes_only)


# ══ G. WIRING + BLAST RADIUS ═════════════════════════════════════════════════════════════════════
section("G. Wiring, route surface and blast radius")

from app.main import app as APP                                                    # noqa: E402
mw = [m.cls.__name__ for m in APP.user_middleware]
ok("G1  BodySizeLimitMiddleware is registered", "BodySizeLimitMiddleware" in mw)
# G1b pinned the middleware stack as an EXACT list of five. Two unrelated middlewares have since
# been registered (AccessLogMiddleware, RateLimitMiddleware), so the literal went red while every
# ORDERING CONSTRAINT it encoded still holds. Re-expressed as the relative order — which is the
# thing that carries meaning — so an unrelated middleware can be added without a false alarm, while
# any actual re-ordering of these five still fails.
_order_ok = all(a in mw and b in mw and mw.index(a) < mw.index(b) for a, b in [
    ("CORSMiddleware", "HardeningMiddleware"),
    ("HardeningMiddleware", "BodySizeLimitMiddleware"),
    ("BodySizeLimitMiddleware", "TenantScopeMiddleware"),
    ("TenantScopeMiddleware", "GZipMiddleware"),
])
ok("G1b relative order holds: CORS → Hardening → BodySizeLimit → TenantScope → GZip", _order_ok, mw)
ok("G1c it is OUTER of TenantScope (an oversized body costs zero identity/DB work)",
   mw.index("BodySizeLimitMiddleware") < mw.index("TenantScopeMiddleware"))
ok("G1d it is INNER of Hardening (a 413 still carries the security headers)",
   mw.index("HardeningMiddleware") < mw.index("BodySizeLimitMiddleware"))

# ── G2-G6b: BLAST-RADIUS checks, which are PR-REVIEW instruments ────────────────────────────────
# Every assertion below diffs the WORKING TREE against BASE_REF (6aadb14) — "this package touched no
# money file, added no migration, added no dependency, moved no route count". They were true and
# valuable while that package was under review. It merged; the branch has moved on by hundreds of
# commits, so today they assert "NOTHING IN THE ENTIRE PRODUCT HAS CHANGED SINCE 6aadb14", which is
# guaranteed false and says nothing whatever about this package. That is not a defect they are
# detecting — it is the instrument outliving the measurement.
#
# NOT DELETED, and deliberately NOT quietly relaxed into passing: they are SKIPPED, counted, and
# printed, so nobody reads this file as "blast radius verified". Re-run them against any base by
# exporting XSS_BASE_REF (e.g. `XSS_BASE_REF=$(git merge-base HEAD main)`), which is what makes them
# useful again the next time this package is genuinely under review.
_BASE = os.environ.get("XSS_BASE_REF")
if not _BASE:
    for _n, _d in [
        ("G2  route count unchanged vs base", "the base's route count is not this branch's"),
        ("G3  NOT ONE money/other-module backend file touched", "diffs the whole tree vs a merged base"),
        ("G3b no other module agent's frontend tree touched", "diffs the whole tree vs a merged base"),
        ("G4  no migration in this package", "diffs the whole tree vs a merged base"),
        ("G5  SEED_VERSION unchanged", "SEED_VERSION has legitimately moved many times since"),
        ("G6  no new backend dependency", "requirements.txt has legitimately grown since"),
        ("G6b no new frontend dependency", "package.json has legitimately grown since"),
    ]:
        skip(_n, f"PR-review blast-radius check against a MERGED base ({BASE_REF}): {_d}. "
                 f"Re-run with XSS_BASE_REF=<ref> to measure a live changeset.")
else:
    _surface = {(pth, m.upper()) for pth, ops in APP.openapi()["paths"].items() for m in ops}
    _pinned_routes = os.environ.get("EXPECT_ROUTES")
    if _pinned_routes:
        ok(f"G2  route surface matches the pinned {_pinned_routes}",
           len(_surface) == int(_pinned_routes), len(_surface))
    else:
        print(f"  --   G2  route surface: {len(_surface)} (path, method) pairs "
              f"— set EXPECT_ROUTES to pin it")
    changed = subprocess.run(["git", "-C", REPO, "diff", "--name-only", _BASE],
                             capture_output=True).stdout.decode().split()
    MONEY_TREES = ("backend/app/modules/commcalc/", "backend/app/modules/asset/",
                   "backend/app/modules/account/", "backend/app/modules/payables/",
                   "backend/app/modules/closing/", "backend/app/modules/storeops/",
                   "backend/app/modules/hr/", "backend/app/modules/storevisit/",
                   "backend/app/modules/billing/")
    touched_money = [f for f in changed if f.startswith(MONEY_TREES)]
    ok("G3  NOT ONE money/other-module backend file touched — no payout number can move",
       not touched_money, touched_money)
    fe_other = [f for f in changed if f.startswith("frontend/src/app/(platform)/")
                and not any(f.startswith("frontend/src/app/(platform)/" + o)
                            for o in ("admin/", "notify/", "helpdesk/", "remediation/",
                                      "configurations/", "failures/", "training/"))]
    ok("G3b no other module agent's frontend tree touched", not fe_other, fe_other)
    ok("G4  no migration in this package (nothing to run in Supabase)",
       not [f for f in changed if f.startswith("database/migrations/")],
       [f for f in changed if f.startswith("database/migrations/")])
    ent = open(os.path.join(_HERE, "app/modules/core/entitlements.py"), encoding="utf-8").read()
    base_ent = subprocess.run(
        ["git", "-C", REPO, "show", f"{_BASE}:backend/app/modules/core/entitlements.py"],
        capture_output=True).stdout.decode()
    sv = re.search(r"SEED_VERSION\s*=\s*(\d+)", ent)
    sv_b = re.search(r"SEED_VERSION\s*=\s*(\d+)", base_ent)
    ok("G5  SEED_VERSION unchanged (this package seeds no new content)",
       bool(sv and sv_b and sv.group(1) == sv_b.group(1)))
    reqs = open(os.path.join(_HERE, "requirements.txt"), encoding="utf-8").read()
    base_reqs = subprocess.run(["git", "-C", REPO, "show", f"{_BASE}:backend/requirements.txt"],
                               capture_output=True).stdout.decode()
    ok("G6  no new backend dependency", reqs == base_reqs)
    pkg = open(os.path.join(REPO, "frontend", "package.json"), encoding="utf-8").read()
    base_pkg = subprocess.run(["git", "-C", REPO, "show", f"{_BASE}:frontend/package.json"],
                              capture_output=True).stdout.decode()
    ok("G6b no new frontend dependency", pkg == base_pkg)

ok("G7  safe_href is importable by any module agent from one place (no re-implementation)",
   os.path.exists(os.path.join(_HERE, "app/modules/core/safe_href.py")))


print(f"\n{'='*96}\nRESULT: {PASS} passed, {FAIL} failed"
      + (f", {len(SKIPPED)} SKIPPED (not passed)" if SKIPPED else ""))
for _s in SKIPPED:
    print(f"  SKIPPED: {_s}")
print('='*96)
sys.exit(1 if FAIL else 0)
