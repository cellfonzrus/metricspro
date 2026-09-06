"""Harness for Carrier Reconciliation (upload & reconcile).

Proves, with NO database and NO network:
  (a) the PURE parser reproduces the totals of a workbook whose every figure this harness controls,
  (b) the Boost-vs-ours diff/merge logic (`build_comparison`) is correct, using Sheet1 (the raw
      per-transaction feed) as a SYNTHETIC "ours" side,
  (c) `parse_payment_type` buckets a line the same way the Boost workbook's column placement does,
  (d) `_recon_ours_paid` buckets, excludes RTR/airtime, and holds ePay == Rebate + Comm.

SELF-CONTAINED (2026-09-06). This harness used to read one specific real customer workbook from an
absolute path under `/root/.claude/uploads/…` — a chat upload that lives outside the repository. That
made it unrunnable anywhere but the one machine that happened to still have the file (it now fails
with FileNotFoundError), and the file itself is a live tenant's carrier money data, which does not
belong in source control. So the fixture is now BUILT IN MEMORY by `build_fixture_workbook()` below,
in the exact layout `carrier_recon.parse_workbook` expects (stacked rebate blocks in cols A-I, the
parallel commissions block in L-P, plus Escalation / Unpaid Devices / Missing / Sheet1).

The authoritative Jul-2026 figures from the original workbook are NOT thrown away: they are kept in
`JUL_2026_AUTHORITATIVE` and asserted verbatim whenever that workbook is actually available — point
the env var `CARRIER_RECON_SAMPLE` at it (or leave the original path in place) and section (a')
runs. With no sample present the harness says so plainly and proves everything else.

Run:  python harness_carrier_recon.py     (from backend/ or from the repo root)
"""

import io
import os
import sys

import openpyxl

# Anchor imports to THIS file's own directory so the harness runs identically from backend/ and from
# the repo root (cf. 564c171f).
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from app.modules.commcalc import carrier_recon as cr  # noqa: E402

# The original chat-upload path, kept only as a fallback location. Override with CARRIER_RECON_SAMPLE.
SAMPLE = os.environ.get("CARRIER_RECON_SAMPLE") or (
    "/root/.claude/uploads/30ae494a-a623-5c2d-9fab-a98a3000e8f7/"
    "4aaa0041-CellularOperationsNJRebate_ReconciliationJuly_2026.xlsx")

# Totals of the real Jul-2026 deliverable, as originally proven. Asserted only when SAMPLE exists.
JUL_2026_AUTHORITATIVE = {
    "store_count": 19, "device_cost": 249382.87, "rebate_expected": 210879.19,
    "rebate_paid": 207179.24, "rebate_diff": -3699.95, "comm_paid": 10533.68,
    "epay_paid": 216646.53, "withhold": 0.0, "escalation_count": 124,
    "escalation_expected": 3822.55, "unpaid_count": 14, "unpaid_cost": 4159.86,
    "missing_count": 1, "raw_txn_count": 1063,
}

CENT = 0.011

_passed = 0
_failed = 0


def check(name, cond, got=None, want=None):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed += 1
        extra = "" if got is None and want is None else f"   (got={got!r} want={want!r})"
        print(f"  FAIL  {name}{extra}")


def approx(a, b, tol=CENT):
    return abs(float(a) - float(b)) <= tol


# ══════════════════════════════════════════════════════════════════════════════════════════════════
# The in-memory fixture workbook — every figure below is chosen here, so the expected totals are
# arithmetic, not folklore. Layout mirrors the Boost/ePay "Rebate Reconciliation" deliverable.
# ══════════════════════════════════════════════════════════════════════════════════════════════════
# store -> (device_cost, cust_paid, financed, rebate_expected, rebate_paid, rebate_diff,
#           rebate_overpaid, gp, comm_paid, comm_gp, withhold)
FIXTURE_STORES = [
    ("100 Main St",  1000.00, 100.00, 200.00,  900.00,  850.00,  -50.00, 0.00, 300.00, 100.00, 50.00,  0.00),
    ("200 Oak Ave",  2000.00, 150.00, 250.00, 1800.00, 1800.00,    0.00, 0.00, 400.00, 200.00, 60.00, 10.00),
    ("300 Pine Rd",  3000.50, 200.00, 300.00, 2500.25, 2400.00, -100.25, 5.00, 500.00, 300.00, 70.00,  0.00),
]
# ePay Paid is the identity ePay == Rebate + Comm, which section (d) also proves on the ours side.
FIXTURE_EPAY = {s[0]: round(s[5] + s[9], 2) for s in FIXTURE_STORES}

# Sheet1 raw feed: per store, two reimbursement-typed lines and two bounty-typed lines, summing back
# to that store's Rebate Paid / Comm Paid exactly. "Device Upgrade Bounty" is deliberately included:
# it reads like a bounty but Boost books it as a REIMBURSEMENT, which is the classifier's hard case.
FIXTURE_FEED = {
    "100 Main St": [("2026 Q3 Promo Upgrade", 400.00, 0.0), ("Device Upgrade Bounty - Month 1", 450.00, 0.0),
                    ("New Activation Bounty - Month 1", 0.0, 60.00), ("Simplified SIM Loading Bounty - Month 1", 0.0, 40.00)],
    "200 Oak Ave": [("2026 Q3 Promo Upgrade", 900.00, 0.0), ("Device Upgrade Bounty - Month 1", 900.00, 0.0),
                    ("New Activation Bounty - Month 1", 0.0, 120.00), ("Simplified SIM Loading Bounty - Month 1", 0.0, 80.00)],
    "300 Pine Rd": [("2026 Q3 Promo Upgrade", 1200.00, 0.0), ("Device Upgrade Bounty - Month 1", 1200.00, 0.0),
                    ("New Activation Bounty - Month 1", 0.0, 180.00), ("Simplified SIM Loading Bounty - Month 1", 0.0, 120.00)],
}
FIXTURE_REPS = {"100 Main St": ["Doe, Jane", "Smith, John"],
                "200 Oak Ave": ["Rivera, Ana", "Chen, Li"],
                "300 Pine Rd": ["Okafor, Ada", "Brown, Sam"]}
FIXTURE_ESCALATIONS = [10.50, 20.25, 30.00]
FIXTURE_UNPAID = [100.00, 59.86]

REB_HEADER = ["Stores", "Total Device Cost", "Customer Paid", "Financed", "Rebate Expected",
              "Rebate Paid", "Rebate Diff", "Rebate Overpaid", "Estimated GP"]
COMM_HEADER = ["Stores", "Total Comm Paid", "Estimated GP", "Withhold", "ePay Paid"]


def _expected_totals():
    """The Grand Total the fixture must produce — summed here rather than typed twice."""
    t = {k: round(sum(s[i] for s in FIXTURE_STORES), 2) for i, k in
         enumerate(["device_cost", "cust_paid", "financed", "rebate_expected", "rebate_paid",
                    "rebate_diff", "rebate_overpaid", "gp", "comm_paid", "comm_gp", "withhold"], start=1)}
    t["epay_paid"] = round(sum(FIXTURE_EPAY.values()), 2)
    t["store_count"] = len(FIXTURE_STORES)
    t["escalation_count"] = len(FIXTURE_ESCALATIONS)
    t["escalation_expected"] = round(sum(FIXTURE_ESCALATIONS), 2)
    t["unpaid_count"] = len(FIXTURE_UNPAID)
    t["unpaid_cost"] = round(sum(FIXTURE_UNPAID), 2)
    t["missing_count"] = 1
    t["raw_txn_count"] = sum(len(v) for v in FIXTURE_FEED.values())
    return t


def build_fixture_workbook() -> bytes:
    """Emit a workbook in the real deliverable's shape. Cols A-I carry three stacked blocks
    (1: per-store rebate summary, 2: a reimbursement breakdown the parser must SKIP, 3: the same
    stores with per-rep rows nested underneath); cols L-P carry the parallel commissions block."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reports"

    def put(row, col, values):
        for i, v in enumerate(values):
            ws.cell(row=row, column=col + i, value=v)

    r = 1
    # ── Block 1 header, plus the commissions header on the SAME row (cols L-P) ──────────────────
    put(r, 1, REB_HEADER)
    put(r, 12, COMM_HEADER)
    r += 1
    for s in FIXTURE_STORES:
        put(r, 1, list(s[:9]))
        put(r, 12, [s[0], s[9], s[10], s[11], FIXTURE_EPAY[s[0]]])
        r += 1
    tt = _expected_totals()
    put(r, 1, ["Grand Total", tt["device_cost"], tt["cust_paid"], tt["financed"],
               tt["rebate_expected"], tt["rebate_paid"], tt["rebate_diff"],
               tt["rebate_overpaid"], tt["gp"]])
    put(r, 12, ["Grand Total", tt["comm_paid"], tt["comm_gp"], tt["withhold"], tt["epay_paid"]])
    r += 2

    # ── Block 2: the reimbursement breakdown. Its header is NOT a rebate header, so the parser must
    #    ignore this block entirely. The absurd figures here exist to make that failure loud. ─────
    put(r, 1, ["Stores", "New Act", "PIC", "Upgrade", "Other"])
    r += 1
    for s in FIXTURE_STORES:
        put(r, 1, [s[0], 999999.0, 999999.0, 999999.0, 999999.0])
        r += 1
    put(r, 1, ["Grand Total", 999999.0, 999999.0, 999999.0, 999999.0])
    r += 2

    # ── Block 3: store rows with per-rep rows nested underneath ─────────────────────────────────
    put(r, 1, REB_HEADER)
    r += 1
    for s in FIXTURE_STORES:
        put(r, 1, list(s[:9]))
        r += 1
        share = round(s[5] / len(FIXTURE_REPS[s[0]]), 2)
        for rep in FIXTURE_REPS[s[0]]:
            put(r, 1, [rep, 0.0, 0.0, 0.0, 0.0, share, 0.0, 0.0, 0.0])
            r += 1
    put(r, 1, ["Grand Total"] + [0.0] * 8)

    # ── Escalation ──────────────────────────────────────────────────────────────────────────────
    esc = wb.create_sheet("Escalation")
    esc.append(["Store", "Trans ID", "Expected Amount", "Note"])
    for i, amt in enumerate(FIXTURE_ESCALATIONS):
        esc.append([FIXTURE_STORES[i % len(FIXTURE_STORES)][0], f"T{i}", amt, "short paid"])

    # ── Unpaid Devices (curated subset of a 60+ column POS export) ───────────────────────────────
    up = wb.create_sheet("Unpaid Devices")
    up.append(["Store", "Sales Rep", "Trans Date Time", "Trans ID", "Device", "Cost", "IMEI"])
    for i, cost in enumerate(FIXTURE_UNPAID):
        up.append([FIXTURE_STORES[i % len(FIXTURE_STORES)][0], "Doe, Jane", "2026-07-14 10:00",
                   f"U{i}", "Phone X", cost, f"35000000000000{i}"])

    # ── Missing ─────────────────────────────────────────────────────────────────────────────────
    ms = wb.create_sheet("Missing")
    ms.append(["Store", "Trans ID", "Note"])
    ms.append([FIXTURE_STORES[0][0], "M1", "not in carrier feed"])

    # ── Sheet1: the raw per-transaction feed ────────────────────────────────────────────────────
    s1 = wb.create_sheet("Sheet1")
    s1.append(["Store", "Payment Type", "Rebate Paid", "Comm Paid", "ePay Paid"])
    for store, lines in FIXTURE_FEED.items():
        for ptype, reb, comm in lines:
            s1.append([store, ptype, reb, comm, round(reb + comm, 2)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sheet1_ours_map(xlsx_bytes):
    """Build a synthetic OUR-side map from Sheet1: {norm(store): {rebate_paid, comm_paid, epay_paid}}
    by summing the raw feed's paid columns per store. This is the ground truth the workbook's Grand
    Total is itself pivoted from, so a correct diff against it is ~0 per store."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb["Sheet1"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    si, ri, ci, ei = (hdr.index("Store"), hdr.index("Rebate Paid"),
                      hdr.index("Comm Paid"), hdr.index("ePay Paid"))

    def f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    out = {}
    for row in it:
        sv = row[si]
        if sv is None or str(sv).strip() == "":
            continue
        k = " ".join(str(sv).strip().split()).lower()
        b = out.setdefault(k, {"rebate_paid": 0.0, "comm_paid": 0.0, "epay_paid": 0.0})
        b["rebate_paid"] = round(b["rebate_paid"] + f(row[ri]), 2)
        b["comm_paid"] = round(b["comm_paid"] + f(row[ci]), 2)
        b["epay_paid"] = round(b["epay_paid"] + f(row[ei]), 2)
    wb.close()
    return out


def section_a_real_sample():
    """The ORIGINAL authoritative Jul-2026 assertions, verbatim — run only when that workbook is
    actually on this machine. Skipped (loudly, with the reason) rather than silently dropped."""
    print("\n(a') Authoritative Jul-2026 totals from the real deliverable:")
    if not os.path.exists(SAMPLE):
        print(f"     SKIPPED — sample workbook not present at {SAMPLE}")
        print("     Reason: it is a one-off chat upload of a live tenant's carrier money data, so it")
        print("     is deliberately NOT committed to the repo. Set CARRIER_RECON_SAMPLE to a copy to")
        print("     run these. Every parser behaviour they cover is also proven above against the")
        print("     in-memory fixture, so nothing here is the harness's only proof of anything.")
        return
    parsed = cr.parse_workbook(open(SAMPLE, "rb").read(), period="Jul-2026")
    t = parsed["totals"]
    A = JUL_2026_AUTHORITATIVE
    for key in ("store_count", "escalation_count", "unpaid_count", "missing_count"):
        check(f"real sample {key} == {A[key]}", t[key] == A[key], t[key], A[key])
    check(f"real sample raw_txn_count == {A['raw_txn_count']}",
          parsed["raw_txn_count"] == A["raw_txn_count"], parsed["raw_txn_count"], A["raw_txn_count"])
    for key in ("device_cost", "rebate_expected", "rebate_paid", "rebate_diff", "comm_paid",
                "epay_paid", "withhold", "escalation_expected", "unpaid_cost"):
        check(f"real sample {key} == {A[key]:,.2f}", approx(t[key], A[key]), t[key], A[key])


def main():
    data = build_fixture_workbook()
    parsed = cr.parse_workbook(data, period="Jul-2026")
    t = parsed["totals"]
    exp = _expected_totals()

    # ── (a) the parser reproduces the fixture's totals ───────────────────────────────────────────
    print("(a) Parser reproduces the fixture workbook's totals:")
    check(f"store count = {exp['store_count']}", t["store_count"] == exp["store_count"],
          t["store_count"], exp["store_count"])
    for key in ("device_cost", "cust_paid", "financed", "rebate_expected", "rebate_paid",
                "rebate_diff", "rebate_overpaid", "gp", "comm_paid", "comm_gp", "withhold",
                "epay_paid"):
        check(f"{key} = {exp[key]:,.2f}", approx(t[key], exp[key]), t[key], exp[key])
    check(f"Escalation items = {exp['escalation_count']}",
          t["escalation_count"] == exp["escalation_count"], t["escalation_count"])
    check(f"Escalation Expected = {exp['escalation_expected']:,.2f}",
          approx(t["escalation_expected"], exp["escalation_expected"]), t["escalation_expected"])
    check(f"Unpaid Devices = {exp['unpaid_count']}", t["unpaid_count"] == exp["unpaid_count"],
          t["unpaid_count"])
    check(f"Unpaid cost = {exp['unpaid_cost']:,.2f}", approx(t["unpaid_cost"], exp["unpaid_cost"]),
          t["unpaid_cost"])
    check(f"Missing = {exp['missing_count']}", t["missing_count"] == exp["missing_count"],
          t["missing_count"])
    check(f"Sheet1 raw_txn_count = {exp['raw_txn_count']}",
          parsed["raw_txn_count"] == exp["raw_txn_count"], parsed["raw_txn_count"])
    check(f"{exp['store_count']} store records parsed",
          len(parsed["stores"]) == exp["store_count"], len(parsed["stores"]))
    check("rep drilldown populated (>0 reps total)",
          sum(len(s["reps"]) for s in parsed["stores"]) > 0)
    check("every rep row is attributed to a store (block 3 nesting)",
          sum(len(s["reps"]) for s in parsed["stores"]) == sum(len(v) for v in FIXTURE_REPS.values()),
          sum(len(s["reps"]) for s in parsed["stores"]))
    check("every store carries an ePay Paid figure",
          all("epay_paid" in s for s in parsed["stores"]))
    # Block 2 is a DIFFERENT breakdown of the same stores; letting it through would corrupt totals.
    check("block 2 (reimbursement breakdown) is ignored, not summed into device_cost",
          approx(t["device_cost"], exp["device_cost"]), t["device_cost"], exp["device_cost"])

    # ── (b) Boost-vs-ours diff/merge with Sheet1 as synthetic ours ───────────────────────────────
    print("\n(b) Boost-vs-ours diff/merge (Sheet1 as synthetic ours):")
    ours = sheet1_ours_map(data)
    # give GP the Boost figure so the money-field diffs are the focus and GP diff is exactly 0 too
    for s in parsed["stores"]:
        k = " ".join(str(s["store"]).strip().split()).lower()
        if k in ours:
            ours[k]["gp"] = round(float(s["gp"]), 2)

    cmp = cr.build_comparison(parsed, ours, resolve=None)
    check(f"all {exp['store_count']} stores matched", all(r["matched"] for r in cmp["per_store"]),
          sum(r["matched"] for r in cmp["per_store"]), exp["store_count"])
    check("no unmatched stores when ours covers all", cmp["unmatched_stores"] == [],
          cmp["unmatched_stores"])

    worst = {f: 0.0 for f in ("rebate_paid", "comm_paid", "epay_paid", "gp")}
    for r in cmp["per_store"]:
        for f in worst:
            worst[f] = max(worst[f], abs(r["diff"][f]))
    check("per-store rebate_paid diff ~ 0 (boost==ours)", worst["rebate_paid"] <= CENT, worst["rebate_paid"])
    check("per-store comm_paid diff ~ 0", worst["comm_paid"] <= CENT, worst["comm_paid"])
    check("per-store epay_paid diff ~ 0", worst["epay_paid"] <= CENT, worst["epay_paid"])
    check("per-store gp diff ~ 0", worst["gp"] <= CENT, worst["gp"])

    # diff == boost - ours identity on a spot store
    spot = cmp["per_store"][0]
    ident = approx(spot["diff"]["rebate_paid"], spot["boost"]["rebate_paid"] - spot["ours"]["rebate_paid"])
    check("diff = boost - ours identity holds", ident)

    # company totals match the parsed Grand Total on the ours side too
    check("ours totals rebate_paid == workbook Grand Total",
          approx(cmp["totals"]["ours"]["rebate_paid"], exp["rebate_paid"]),
          cmp["totals"]["ours"]["rebate_paid"], exp["rebate_paid"])
    check("ours totals epay_paid == workbook Grand Total",
          approx(cmp["totals"]["ours"]["epay_paid"], exp["epay_paid"]),
          cmp["totals"]["ours"]["epay_paid"], exp["epay_paid"])
    check("boost totals comm_paid == workbook Grand Total",
          approx(cmp["totals"]["boost"]["comm_paid"], exp["comm_paid"]),
          cmp["totals"]["boost"]["comm_paid"], exp["comm_paid"])

    # ── unmatched handling: drop one store from ours ─────────────────────────────────────────────
    print("\n(b') Unmatched handling (drop one store from ours):")
    dropped_name = parsed["stores"][0]["store"]
    dropped_key = " ".join(str(dropped_name).strip().split()).lower()
    ours2 = {k: v for k, v in ours.items() if k != dropped_key}
    cmp2 = cr.build_comparison(parsed, ours2, resolve=None)
    check("dropped store is listed in unmatched_stores", dropped_name in cmp2["unmatched_stores"],
          cmp2["unmatched_stores"])
    check("exactly one unmatched store", len(cmp2["unmatched_stores"]) == 1,
          len(cmp2["unmatched_stores"]))
    drow = next(r for r in cmp2["per_store"] if r["store"] == dropped_name)
    check("dropped store not matched", drow["matched"] is False, drow["matched"])
    check("dropped store ours zeroed", drow["ours"]["rebate_paid"] == 0.0, drow["ours"]["rebate_paid"])
    check("dropped store diff == boost (nothing on ours side)",
          approx(drow["diff"]["rebate_paid"], drow["boost"]["rebate_paid"]))
    check("dropped store STILL returned (never silently dropped)",
          any(r["store"] == dropped_name for r in cmp2["per_store"]))

    # ── (c) classifier reproduces Boost's Rebate/Comm split on Sheet1's Payment Type column ───────
    print("\n(c) parse_payment_type bucketing agrees with Boost's Rebate/Comm column placement:")
    from app.modules.commcalc.discrepancy_engine import parse_payment_type
    from app.modules.commcalc.router import (_RECON_BOUNTY_COMPS, _RECON_REIMB_COMPS)

    def _bucket(payment_type):
        """Mirror `_recon_payment_bucketer` with NO org payment_categories config (comp_type only)."""
        try:
            comp = parse_payment_type(payment_type)[0]
        except Exception:
            comp = ""
        if comp in _RECON_REIMB_COMPS:
            return "reimbursement"
        if comp in _RECON_BOUNTY_COMPS:
            return "bounty"
        return "other"

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb["Sheet1"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ci, ri, pti = hdr.index("Comm Paid"), hdr.index("Rebate Paid"), hdr.index("Payment Type")

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    # Ground truth = which Boost column carries the money. Test rows that carry a Payment Type AND have
    # money in EXACTLY ONE of {Comm Paid, Rebate Paid} (a row split across both is a Boost-side judgement
    # call, not a classifier fact). Expected: comm-only → bounty, rebate-only → reimbursement.
    tally = {"bounty": [0, 0], "reimbursement": [0, 0]}
    mism = []
    for row in it:
        pt = row[pti]
        if pt is None or str(pt).strip() == "":
            continue
        comm, reb = _f(row[ci]), _f(row[ri])
        if (comm > 0) == (reb > 0):        # need money in exactly one column
            continue
        expected = "bounty" if comm > 0 else "reimbursement"
        got = _bucket(pt)
        tally[expected][1] += 1
        if got == expected:
            tally[expected][0] += 1
        else:
            mism.append((expected, got, comm, reb, str(pt)[:48]))
    wb.close()

    tot_a = sum(v[0] for v in tally.values())
    tot_t = sum(v[1] for v in tally.values())
    for b in ("reimbursement", "bounty"):
        a, t2 = tally[b]
        rate = (100.0 * a / t2) if t2 else 0.0
        print(f"     {b:14s} agreement {a}/{t2}  ({rate:.0f}%)")
    print(f"     overall        agreement {tot_a}/{tot_t}  "
          f"({(100.0 * tot_a / tot_t) if tot_t else 0.0:.0f}%)")
    if mism:
        print("     residual mismatches (Boost split a promo line into Comm — a Boost-side quirk, "
              "not a classifier error):")
        for e, g, c, r, p in mism:
            print(f"       exp {e} got {g}  comm={c:.2f} reb={r:.2f}  {p}")
    check("reimbursement bucket agreement == 100%", tally["reimbursement"][1] > 0 and
          tally["reimbursement"][0] == tally["reimbursement"][1],
          f"{tally['reimbursement'][0]}/{tally['reimbursement'][1]}")
    check("overall classification agreement >= 90%", tot_t > 0 and tot_a / tot_t >= 0.90,
          f"{tot_a}/{tot_t}")

    # ── (d) ours-side helper: bucketing + the ePay == Rebate + Comm identity (real `_recon_ours_paid`) ─
    print("\n(d) _recon_ours_paid: bucketing, RTR exclusion, and ePay == Rebate + Comm identity:")
    import types as _types
    from app.modules.commcalc.router import _recon_ours_paid

    class _FakeQ:
        def __init__(self, rows):
            self._rows = rows

        def select(self, *a, **k):
            return self

        def eq(self, *a, **k):
            return self

        def in_(self, *a, **k):
            return self

        def limit(self, *a, **k):
            return self

        def execute(self):
            return _types.SimpleNamespace(data=self._rows)

    class _FakeClient:
        def __init__(self, tables):
            self._tables = tables

        def schema(self, _s):
            return self

        def table(self, name):
            return _FakeQ(self._tables.get(name, []))

    pay_rows = [
        {"business_address": "100 Main St", "payment_type": "3187138: New Activation Bounty - Month 1",
         "amount": 5.0},                                                  # bounty
        {"business_address": "100 Main St", "payment_type": "2026 Q3 Promo Upgrade",
         "amount": 50.0},                                                 # reimbursement
        {"business_address": "100 Main St", "payment_type": "Device Upgrade Bounty - Month 1",
         "amount": 500.0},                                                # reimbursement (DUPGB)
        {"business_address": "100 Main St", "payment_type": "Auto Top Up Monthly Incentive",
         "amount": 999.0},                                                # OTHER / RTR — excluded
        {"business_address": "200 Oak Ave", "payment_type": "Simplified SIM Loading Bounty - Month 1",
         "amount": 3.0},                                                  # bounty
        {"business_address": "200 Oak Ave", "payment_type": "SIM Card Reimbursement",
         "amount": 2.5},                                                  # reimbursement (SIMCR)
    ]
    fake = _FakeClient({"payment_categories": [], "raw_payment_detail": pay_rows})
    into, hnotes = {}, []
    _recon_ours_paid(fake, "org-test", "Jul-2026", lambda a: (a, None), into, hnotes)
    a = into.get("100 main st", {})
    b = into.get("200 oak ave", {})
    check("store A comm_paid == 5.00 (NAB bounty)", approx(a.get("comm_paid", 0), 5.0), a.get("comm_paid"))
    check("store A rebate_paid == 550.00 (promo upgrade + device upgrade)",
          approx(a.get("rebate_paid", 0), 550.0), a.get("rebate_paid"))
    check("store A ePay excludes the $999 RTR/airtime line",
          approx(a.get("epay_paid", 0), 555.0), a.get("epay_paid"))
    check("store A identity ePay == Rebate + Comm",
          approx(a.get("epay_paid", 0), a.get("rebate_paid", 0) + a.get("comm_paid", 0)))
    check("store B comm_paid == 3.00 / rebate_paid == 2.50",
          approx(b.get("comm_paid", 0), 3.0) and approx(b.get("rebate_paid", 0), 2.5),
          (b.get("comm_paid"), b.get("rebate_paid")))
    check("store B identity ePay == Rebate + Comm",
          approx(b.get("epay_paid", 0), b.get("rebate_paid", 0) + b.get("comm_paid", 0)))

    # ── (a') the real deliverable, when it is available ─────────────────────────────────────────
    section_a_real_sample()

    print(f"\n==== {_passed} passed, {_failed} failed ====")
    return 1 if _failed else 0


if __name__ == "__main__":
    sys.exit(main())
