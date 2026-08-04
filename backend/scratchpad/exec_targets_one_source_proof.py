"""Proof for agent/commission/exec-targets-one-source — the owner's core requirement (2026-07-16):
over IDENTICAL inputs, Executive MTD's cumulative figures EQUAL the Sales Report's aggregation, and Daily
Targets' actuals come from that SAME aggregation. No hand-waving.

Drives the REAL router functions over an in-memory FakeClient (house style; no DB/network):
  (a) a synthetic open month where the feed and raw_sales DISAGREE per (day × store) cell →
      _sales_cell_agg rollup == Sales Report totals == Exec MTD cumulative == Targets actuals, for
      EVERY shared bucket (txns / activations / byod / upgrades / swaps / revenue / gp / accessory$).
      Includes a multi-line AAL transaction so distinct-trans_id counting is actually exercised.
  (b) the MTD date-cut: a line dated AFTER the cut appears in the Sales Report's monthly total but NOT
      in Exec MTD's cumulative — the ONE intentional difference; with today past that date they re-equal.
  (c) voided / Return / blank-or-'admin' rep rows are excluded IDENTICALLY by all three surfaces.
  (d) the REAL luxelink Total-Wireless sample driven through all three shared paths → they agree.

Run:  cd backend && python3 scratchpad/exec_targets_one_source_proof.py

UPDATED 2026-07-18 (agent/commission/sales-capture-fix): `_sales_rows_union` gained a dedup-by-trans_id
COMPLETENESS backfill (luxelink '957 Pennsylvania Ave' undercount). The winner-take-all cell-grain merge
used to DROP a raw_sales transaction whose trans_id the feed lacked on a feed-led store-day — but that IS a
real sale the incomplete hourly feed missed (the owner's bug). The fix keeps the feed's copy of SHARED
trans_ids (a STALE duplicate like 'A' below, same trans_id, is still dropped) and only surfaces raw-ONLY
trans_ids ('Z' below). The all-three-surfaces-agree invariant is UNCHANGED (they call the same union); the
ground-truth values here rise by exactly the recovered raw-only transaction. See the updated (a0)/(a1)/(b).
"""


def run_route(x):
    """Call a commcalc route handler in EITHER shape.

    ASYNC-SWEEP 2026-08-04: commcalc's zero-`await` route handlers were converted from `async def` to
    `def` so FastAPI runs them in the threadpool instead of on the single uvicorn event loop (an
    `async def` doing blocking Supabase I/O froze the whole product for its duration). The only textual
    change was the keyword. This helper awaits a coroutine when it gets one and passes a plain result
    straight through, so the proof works against BOTH shapes and needs no further edit if a handler
    ever legitimately becomes a coroutine again."""
    import asyncio as _a
    return _a.run(x) if _a.iscoroutine(x) else x
import os, sys, asyncio, calendar as _cal
from datetime import date as _date
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import app.modules.commcalc.router as R
from app.modules.commcalc.calculator import classify_contract_type as CCT

PASS = 0
FAIL = 0


def check(name, cond, extra=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ok  {name}")
    else:
        FAIL += 1
        print(f"FAIL  {name}   {extra}")


# ── in-memory fake supabase client (honours eq / in_ / neq / count) ──────────────────────────────
class FakeResult:
    def __init__(self, data=None, count=None):
        self.data = data or []
        self.count = count


class FakeQuery:
    def __init__(self, store, table):
        self.store = store
        self.t = table
        self.f = []
        self.cnt = False
        self.rng = None

    def select(self, *a, **k):
        if k.get('count') == 'exact':
            self.cnt = True
        return self

    def eq(self, c, v):
        self.f.append(('eq', c, v))
        return self

    def neq(self, c, v):
        self.f.append(('neq', c, v))
        return self

    def in_(self, c, v):
        self.f.append(('in', c, list(v)))
        return self

    def limit(self, n):
        return self

    def range(self, a, b):
        self.rng = (a, b)
        return self

    def order(self, *a, **k):
        return self

    def _m(self, r):
        for k, c, v in self.f:
            rv = r.get(c)
            if k == 'eq' and rv != v:
                return False
            if k == 'neq' and rv == v:
                return False
            if k == 'in' and rv not in v:
                return False
        return True

    def execute(self):
        rows = self.store.setdefault(self.t, [])
        m = [r for r in rows if self._m(r)]
        if self.rng:
            a, b = self.rng
            m = m[a:b + 1]
        if self.cnt:
            return FakeResult(data=m, count=len(m))
        return FakeResult(data=[dict(r) for r in m])


class FakeSchema:
    def __init__(self, store):
        self.store = store

    def table(self, t):
        return FakeQuery(self.store, t)

    def rpc(self, *a, **k):
        raise Exception('no rpc in this proof')


class FakeClient:
    def __init__(self, store):
        self.store = store

    def schema(self, s):
        return FakeSchema(self.store)


ORG = 'o'
_T = _date.today()
OPEN = f"{_T.year}-{_T.month:02d}"                       # the current (open) month
MONTHNAME = f"{_cal.month_name[_T.month]} {_T.year}"     # raw_sales spelling (period-spelling agnosticism)
DIM = _cal.monthrange(_T.year, _T.month)[1]
CUT = _date(_T.year, _T.month, 15)                       # injected "today" → MTD cut at the 15th
LATE = _date(_T.year, _T.month, min(20, DIM))            # a day AFTER the cut, still in the month
D10 = f"{OPEN}-10"
DLATE = f"{OPEN}-{min(20, DIM):02d}"


def row(store, rep, tid, ct, day=D10, cat='CellPhone', dept='', ext=100.0, gp=20.0, pdesc='',
        voided='', tt='', period=OPEN):
    return {'org_id': ORG, 'period': period, 'trans_id': tid, 'trans_date': day, 'store': store,
            'salesperson': rep, 'user_login': (rep or '').lower(), 'category': cat, 'department': dept,
            'contract_type': ct, 'product_desc': pdesc, 'ext_price': ext, 'gp': gp,
            'voided': voided, 'trans_type': tt}


# ══ (a) feed vs raw DISAGREE per cell — the cell-grain union resolves them; all 3 surfaces then agree ══
# FEED store S1/day10 (feed WINS this cell): A = a 2-line AAL activation (distinct-txn = 1), P = a Port
# (premium+port), B = BYOD, U = Upgrade, SW = a SIM Swap (classifier None → swaps only), ACC = accessory.
# Plus 3 rows the skip rules must drop: V (voided), RET (Return), ADM ('admin' rep).
feed = [
    row('S1', 'ALICE', 'A', 'Activation AAL', ext=100.0),
    row('S1', 'ALICE', 'A', 'Activation AAL', ext=100.0),   # 2nd line, same trans_id (AAL) → still 1 txn
    row('S1', 'ALICE', 'P', 'Port-In', ext=100.0),          # premium AND a Port (sub-split of premium)
    row('S1', 'BOB', 'B', 'BYOD Activation', ext=100.0),
    row('S1', 'BOB', 'U', 'Upgrade', ext=100.0),
    row('S1', 'CARLA', 'SW', 'SIM Swap', ext=50.0),
    row('S1', 'CARLA', 'ACC', '', cat='Accessory', ext=40.0),
    row('S1', 'ALICE', 'V', 'Activation', ext=100.0, voided='Yes'),      # voided → dropped
    row('S1', 'ALICE', 'RET', 'Activation', ext=100.0, tt='Return'),     # Return → dropped
    row('S1', 'admin', 'ADM', 'Activation', ext=100.0),                  # 'admin' rep → dropped
]
# RAW store S1/day10 (feed leads the cell): 'A' shares the feed's trans_id → its STALE raw copy is
# DROPPED by the trans_id dedup (feed wins shared trans_ids); 'Z' is a raw-ONLY trans_id the feed missed
# → the completeness backfill now SURFACES it (a real sale, the luxelink undercount fix). + RAW store
# S2/day10 (feed has NO S2 rows → this cell is FILLED from raw): D = Upgrade, E = accessory.
raw = [
    row('S1', 'ALICE', 'A', 'Activation', ext=999.0, period=MONTHNAME),   # stale dup of A (shared id) → dropped
    row('S1', 'ZED', 'Z', 'Activation', ext=999.0, period=MONTHNAME),     # raw-only trans_id → now backfilled
    row('S2', 'DAN', 'D', 'Upgrade', ext=100.0, period=MONTHNAME),
    row('S2', 'DAN', 'E', '', cat='Accessory', ext=60.0, period=MONTHNAME),
]
acc_cfg = [{'org_id': ORG, 'departments': [], 'categories': ['accessory'], 'product_keywords': [],
            'acima_tenders': []}]
store = {'daily_sales_feed': [dict(r) for r in feed], 'raw_sales': [dict(r) for r in raw],
         'accessory_config': acc_cfg, 'store_mapping': [], 'stores': [], 'exec_metric_config': []}
c = FakeClient(store)

# Ground truth: the resolved union, aggregated by the ONE shared function.
urows, umeta = R._sales_rows_union(c, ORG, OPEN, cols=R._SALES_DISPLAY_COLS)
acfg = R._accessory_config(c, ORG)
cells = R._sales_cell_agg(urows, acfg)
GT = {
    'txns': sum(len(a['_txn']) for a in cells.values()),
    'activations': sum(len(a['_prem']) for a in cells.values()),
    'byod': sum(len(a['_byod']) for a in cells.values()),
    'upgrades': sum(len(a['_upg']) for a in cells.values()),
    'swaps': sum(len(a['_swap']) for a in cells.values()),
    'accessory_rev': round(sum(a['accessory_rev'] for a in cells.values()), 2),
    'revenue': round(sum(a['revenue'] for a in cells.values()), 2),
    'gp': round(sum(a['gp'] for a in cells.values()), 2),
}

print("(a0) the cell-grain union resolved the feed↔raw disagreement (feed wins S1, raw fills S2)")
shown_stores = {r['store'] for r in urows}
check("union shows BOTH stores (S1 feed-led, S2 filled from raw)", shown_stores == {'S1', 'S2'}, shown_stores)
check("S1's STALE raw dup of A (shared trans_id) was DROPPED (feed wins shared ids)",
      not any(r['trans_id'] == 'A' and r['ext_price'] == 999.0 for r in urows))
check("S1's raw-ONLY Z (feed missed it) is now BACKFILLED (completeness fix)",
      any(r['trans_id'] == 'Z' for r in urows))
check("S2 (feed-less) filled from raw_sales (D + E present)",
      {r['trans_id'] for r in urows if r['store'] == 'S2'} == {'D', 'E'})

print("\n(a1) INDEPENDENT hand model of the resolved union == the shared _sales_cell_agg rollup")
# valid txns: A(prem), P(prem+port), B(byod), U(upg), SW(swap-only), ACC(acc), Z(prem, raw-only backfill);
# S2: D(upg), E(acc). Z adds one activation/txn + its ext 999 / gp 20 (the recovered raw-only sale).
HAND = {'txns': 9, 'activations': 3, 'byod': 1, 'upgrades': 2, 'swaps': 1,
        'accessory_rev': 100.0, 'revenue': 1749.0, 'gp': 200.0}
for k, v in HAND.items():
    check(f"ground-truth {k} == {v}", GT[k] == v, f"got {GT[k]}")

print("\n(a2) Sales Report totals == ground truth (same shared aggregation)")
_orig = R.sb
R.sb = lambda: c
try:
    sr = run_route(R.sales_report(period=OPEN, authorization="", org_id=ORG))
finally:
    R.sb = _orig
srt = sr['totals']
for k in HAND:
    check(f"Sales Report totals.{k} == ground truth", srt[k] == GT[k], f"{srt[k]} vs {GT[k]}")

print("\n(a3) Exec MTD cumulative (today past all rows) == Sales Report, bucket-for-bucket")
ex = R._exec_mtd(c, ORG, OPEN, today=_date(_T.year, _T.month, DIM))   # cut = month end → no rows dropped
et = ex['by_location']['total']
check("Exec (activation + port) == Sales Report activations",
      et['activation'] + et['port'] == srt['activations'], f"{et['activation']}+{et['port']} vs {srt['activations']}")
check("Exec byod == Sales Report byod", et['byod'] == srt['byod'], f"{et['byod']} vs {srt['byod']}")
check("Exec upgrade == Sales Report upgrades", et['upgrade'] == srt['upgrades'], f"{et['upgrade']} vs {srt['upgrades']}")
check("Exec total_activation == Sales Report (activations+byod+upgrades)",
      et['total_activation'] == srt['activations'] + srt['byod'] + srt['upgrades'],
      f"{et['total_activation']} vs {srt['activations'] + srt['byod'] + srt['upgrades']}")
check("Exec acc_sales == Sales Report accessory_rev", round(et['acc_sales'], 2) == srt['accessory_rev'],
      f"{et['acc_sales']} vs {srt['accessory_rev']}")
# by_location total and by_employee total reconcile (same cells, two groupings)
ee = ex['by_employee']['total']
check("Exec by_location total == by_employee total (total_activation)", et['total_activation'] == ee['total_activation'])

print("\n(a4) Daily Targets actuals (_fetch_actuals) == Sales Report, bucket-for-bucket")
R.sb = lambda: c
try:
    fa = R._fetch_actuals(c, ORG, OPEN)
finally:
    R.sb = _orig
t_prem = sum(a['prem_count'] for a in fa)
t_byod = sum(a['byod_count'] for a in fa)
t_upg = sum(a['upg_count'] for a in fa)
t_acc = round(sum(a['acc_gp'] for a in fa), 2)
check("Targets prem_count == Sales Report activations", t_prem == srt['activations'], f"{t_prem} vs {srt['activations']}")
check("Targets byod_count == Sales Report byod", t_byod == srt['byod'], f"{t_byod} vs {srt['byod']}")
check("Targets upg_count == Sales Report upgrades", t_upg == srt['upgrades'], f"{t_upg} vs {srt['upgrades']}")
check("Targets acc_gp == Sales Report accessory_rev", t_acc == srt['accessory_rev'], f"{t_acc} vs {srt['accessory_rev']}")

# ══ (b) MTD date-cut — a row AFTER the cut is in the Sales Report monthly total, not in Exec MTD ══
print("\n(b) MTD date-cut applied to the SAME rows (the ONE intentional difference)")
store_b = {k: [dict(r) for r in v] for k, v in store.items()}
store_b['daily_sales_feed'].append(row('S1', 'ALICE', 'FUT', 'Activation', day=DLATE, ext=100.0))  # after cut
cb = FakeClient(store_b)
R.sb = lambda: cb
try:
    sr_b = run_route(R.sales_report(period=OPEN, authorization="", org_id=ORG))
finally:
    R.sb = _orig
ex_cut = R._exec_mtd(cb, ORG, OPEN, today=CUT)                       # cut at the 15th → FUT (20th) excluded
ex_full = R._exec_mtd(cb, ORG, OPEN, today=_date(_T.year, _T.month, DIM))  # month-end → FUT included
sr_act = sr_b['totals']['activations']
cut_act = ex_cut['by_location']['total']['activation'] + ex_cut['by_location']['total']['port']
full_act = ex_full['by_location']['total']['activation'] + ex_full['by_location']['total']['port']
# base activations now 3 (A, P + the backfilled raw-only Z, all on day10) + FUT (day20) = 4 with no cut.
check("Sales Report (no cut) counts the post-cut FUT row (activations == 4)", sr_act == 4, f"got {sr_act}")
check("Exec MTD (cut=15th) EXCLUDES the day-20 FUT row (activations == 3)", cut_act == 3, f"got {cut_act}")
check("difference is EXACTLY the one post-cut activation", sr_act - cut_act == 1)
check("Exec MTD (today=month-end) re-includes FUT → equals Sales Report (3)", full_act == sr_act, f"{full_act} vs {sr_act}")

# ══ (c) voided / Return / admin excluded IDENTICALLY by all three ══
print("\n(c) voided / Return / blank-or-admin rep excluded identically across the three surfaces")
# Build a clean twin WITHOUT the 3 skip rows; every surface's totals must be byte-identical to the dirty set.
clean = [r for r in feed if r['trans_id'] not in ('V', 'RET', 'ADM')]
store_c = {**{k: [dict(r) for r in v] for k, v in store.items()}, 'daily_sales_feed': [dict(r) for r in clean]}
cc = FakeClient(store_c)
R.sb = lambda: cc
try:
    sr_c = run_route(R.sales_report(period=OPEN, authorization="", org_id=ORG))
finally:
    R.sb = _orig
check("Sales Report totals identical with vs without the voided/Return/admin rows",
      sr_c['totals'] == sr['totals'], f"{sr_c['totals']} vs {sr['totals']}")
ex_c = R._exec_mtd(cc, ORG, OPEN, today=_date(_T.year, _T.month, DIM))
check("Exec MTD total_activation identical with vs without the skip rows",
      ex_c['by_location']['total']['total_activation'] == et['total_activation'])
R.sb = lambda: cc
try:
    fa_c = R._fetch_actuals(cc, ORG, OPEN)
finally:
    R.sb = _orig
check("Targets prem_count identical with vs without the skip rows",
      sum(a['prem_count'] for a in fa_c) == t_prem)

# ══ (d) the REAL luxelink sample through all three shared paths → they agree ══
print("\n(d) REAL luxelink sample — Sales Report == Exec MTD == Targets on the shared buckets")
import openpyxl
SAMPLE = "/workspaces/commcalc/commcalc/My Sales Transaction Details Legacy New with all columns (3).xlsx"
wb = openpyxl.load_workbook(SAMPLE, data_only=True)
ws = wb[wb.sheetnames[0]]
xr = list(ws.iter_rows(values_only=True))
hdr = list(xr[0])
IDX = {h: i for i, h in enumerate(hdr)}


def g(r, n):
    v = r[IDX[n]]
    return '' if v is None else v


def sfl(x):
    try:
        return float(x)
    except Exception:
        return 0.0


st_name = None
tid = None
sfeed = []
for r in xr[1:]:
    s0 = str(r[0]).strip() if r[0] is not None else ''
    if s0.startswith('Store:'):
        st_name = s0.split(':', 1)[1].strip()
        continue
    if s0.startswith('Trans ID:'):
        tid = s0.split(':', 1)[1].strip()
        continue
    if r[0] is None:
        continue
    sfeed.append({'org_id': 'LUX', 'period': OPEN, 'store': st_name, 'salesperson': g(r, 'Salesperson'),
                  'department': str(g(r, 'Department')).strip(), 'category': str(g(r, 'Category')).strip(),
                  'product_desc': str(g(r, 'Product Desc')).strip(), 'contract_type': str(g(r, 'Contract Type')).strip(),
                  'ext_price': sfl(g(r, 'Ext Price')), 'gp': sfl(g(r, 'GP')),
                  'voided': str(g(r, 'Voided')).strip(), 'trans_type': str(g(r, 'Trans Type')).strip(),
                  'user_login': str(g(r, 'Salesperson')).strip().lower(),
                  'trans_id': tid, 'trans_date': str(g(r, 'Trans Date Time'))[:10]})
lux_acc = [{'org_id': 'LUX', 'departments': [], 'categories': ['accessory', 'handsetbranded', 'accessories'],
            'product_keywords': [], 'acima_tenders': []}]
store_d = {'daily_sales_feed': [dict(r) for r in sfeed], 'raw_sales': [], 'accessory_config': lux_acc,
           'store_mapping': [], 'stores': [], 'exec_metric_config': []}
cd = FakeClient(store_d)
R.sb = lambda: cd
try:
    sr_d = run_route(R.sales_report(period=OPEN, authorization="", org_id='LUX'))
finally:
    R.sb = _orig
ex_d = R._exec_mtd(cd, 'LUX', OPEN, today=_date(_T.year, _T.month, DIM))
R.sb = lambda: cd
try:
    fa_d = R._fetch_actuals(cd, 'LUX', OPEN)
finally:
    R.sb = _orig
srd = sr_d['totals']
exd = ex_d['by_location']['total']
check("luxelink: Exec (activation+port) == Sales Report activations",
      exd['activation'] + exd['port'] == srd['activations'], f"{exd['activation']}+{exd['port']} vs {srd['activations']}")
check("luxelink: Exec byod == Sales Report byod", exd['byod'] == srd['byod'], f"{exd['byod']} vs {srd['byod']}")
check("luxelink: Exec upgrade == Sales Report upgrades", exd['upgrade'] == srd['upgrades'], f"{exd['upgrade']} vs {srd['upgrades']}")
check("luxelink: Exec acc_sales == Sales Report accessory_rev",
      round(exd['acc_sales'], 2) == srd['accessory_rev'], f"{exd['acc_sales']} vs {srd['accessory_rev']}")
check("luxelink: Targets prem_count == Sales Report activations",
      sum(a['prem_count'] for a in fa_d) == srd['activations'])
check("luxelink: Targets byod_count == Sales Report byod", sum(a['byod_count'] for a in fa_d) == srd['byod'])
check("luxelink: Targets upg_count == Sales Report upgrades", sum(a['upg_count'] for a in fa_d) == srd['upgrades'])
check("luxelink: Targets acc_gp == Sales Report accessory_rev",
      round(sum(a['acc_gp'] for a in fa_d), 2) == srd['accessory_rev'])
# port-idv (owner ruling 2026-07-16): 'Port with IDV' now classifies 'premium' -> +5 distinct-txn
# activations over this sample (14 -> 19). The Targets/Exec/Sales-Report EQUALITY checks above still
# hold (all three consume the shared classify_contract_type); only this hardcoded literal moves.
check("luxelink: Sales Report activations == 19 distinct-txn (14 + 5 Port-with-IDV; owner 2026-07-16)", srd['activations'] == 19, f"got {srd['activations']}")

print(f"\n=================  {PASS} passed, {FAIL} failed  =================")
sys.exit(1 if FAIL else 0)
