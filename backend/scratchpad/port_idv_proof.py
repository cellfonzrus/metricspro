"""
port-idv proof — "Port with IDV" IS an activation (owner ruling 2026-07-16).

Proves, with NO live DB (Supabase is web-only):
  (a) every contract-type label that does NOT contain 'idv' classifies IDENTICALLY to the pre-change
      classifier (byte-for-byte over a broad corpus + the REAL luxelink sample's distinct labels);
  (b) the ONLY delta introduced is: labels that contain 'idv' AND were None before now classify as
      'premium' (IDV labels that were already premium/byod/upgrade are unchanged — byod/upgrade first);
  (c) the owner-visible number delta over the REAL luxelink July sample, driven through the ACTUAL
      _exec_mtd + _sales_cell_agg engine (not a re-implementation).

Run:  cd backend && python3 scratchpad/port_idv_proof.py
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app.modules.commcalc.calculator import classify_contract_type as CC  # NEW (post-change)

PASS = FAIL = 0
def check(name, cond, extra=''):
    global PASS, FAIL
    if cond:
        PASS += 1
        # print(f"  ok  {name}")
    else:
        FAIL += 1
        print(f"  FAIL {name}  {extra}")

# ── Frozen copy of the OLD classifier (pre-change), to diff against ──────────────────────────────
_OLD_PREMIUM_ACT = {
    'Activation', 'Port-In', 'Add A Line', 'Port-In Add A Line',
    'Eligible Port-In Activation', 'Activation Add A Line', 'Eligible Port-In Add A Line',
}
_OLD_KEYS = ('activation', 'port-in', 'port in', 'add a line', 'add-a-line', 'new line', ' aal', 'aal ')
def CC_OLD(ct):
    c = (ct or '').strip()
    if not c:
        return None
    cl = c.lower()
    if 'byod' in cl:
        return 'byod'
    if 'upgrade' in cl:
        return 'upgrade'
    if c in _OLD_PREMIUM_ACT or any(k in cl for k in _OLD_KEYS):
        return 'premium'
    return None

# ── Corpus: real sample labels + drift + known B2B sets + adversarial ────────────────────────────
CORPUS = [
    '', 'Activation', 'Activation AAL', 'Activation With IDV', 'BYOD Activation', 'BYOD Port',
    'BYOD Port AAL', 'Port with IDV', 'Port with IDV AAL', 'Upgrade',
    # IDV drift (slash / casing / spacing / order)
    'Port w/ IDV', 'PORT WITH IDV', 'port with idv', 'Port With IDV', 'Port w IDV', 'IDV Port',
    'Activation IDV', 'BYOD Port with IDV', 'Upgrade with IDV',
    # known B2B set members
    'BYOD', 'BYOD Port-In', 'BYOD Add A Line', 'BYOD Swap', 'Upgrade Port-In', 'Device Upgrade',
    'Port-In', 'Add A Line', 'Port-In Add A Line', 'Eligible Port-In Activation', 'New Activation',
    'Standard Activation', 'New Line',
    # adversarial non-activation
    'Accessory', 'SIM Kit', 'Bill Payment', 'Return', 'Trade In',
]

print("== A. (a) non-IDV labels byte-identical  &  (b) only delta = IDV None->premium ==")
delta_labels = []
for lab in CORPUS:
    old, new = CC_OLD(lab), CC(lab)
    if 'idv' not in lab.lower():
        # (a) — nothing without 'idv' may move
        check(f"non-idv unchanged: {lab!r} old={old!r} new={new!r}", new == old, f"MOVED to {new!r}")
    else:
        if old is None:
            # (b) — an IDV label that was None must now be premium (it is not byod/upgrade)
            check(f"idv None->premium: {lab!r}", new == 'premium', f"got {new!r}")
            delta_labels.append(lab)
        else:
            # an IDV label already classified (premium/byod/upgrade) must be UNCHANGED (ordering intact)
            check(f"idv already-classified unchanged: {lab!r} ({old!r})", new == old, f"MOVED to {new!r}")

# the delta set is EXACTLY the IDV-was-None labels, and every one goes to 'premium'
moved = [l for l in CORPUS if CC(l) != CC_OLD(l)]
check("moved set == IDV-was-None set", set(moved) == set(delta_labels), f"moved={moved}")
check("every moved label goes None->premium",
      all(CC_OLD(l) is None and CC(l) == 'premium' for l in moved), f"moved={moved}")
# ordering guarantees the task called out
check("BYOD Port with IDV stays byod (byod check first)", CC('BYOD Port with IDV') == 'byod')
check("Upgrade with IDV stays upgrade (upgrade check first)", CC('Upgrade with IDV') == 'upgrade')
check("bare 'port' NOT reclassified (no lone-port key added)", CC('Port Out') is None and CC_OLD('Port Out') is None)

# ── B. owner-visible delta over the REAL luxelink July sample, via the ACTUAL engine ─────────────
print("\n== B. real-sample delta through the ACTUAL _exec_mtd / _sales_cell_agg engine ==")
SAMPLE = "/workspaces/commcalc/commcalc/My Sales Transaction Details Legacy New with all columns (3).xlsx"
if os.path.exists(SAMPLE):
    import openpyxl
    import app.modules.commcalc.router as R
    wb = openpyxl.load_workbook(SAMPLE, data_only=True); ws = wb[wb.sheetnames[0]]
    xr = list(ws.iter_rows(values_only=True)); hdr = list(xr[0]); IDX = {h: i for i, h in enumerate(hdr)}
    def g(r, n):
        v = r[IDX[n]]; return '' if v is None else v
    def sf(x):
        try: return float(x)
        except Exception: return 0.0
    def build_feed(cat):
        store = tid = None; out = []
        for r in xr[1:]:
            s0 = str(r[0]).strip() if r[0] is not None else ''
            if s0.startswith('Store:'): store = s0.split(':', 1)[1].strip(); continue
            if s0.startswith('Trans ID:'): tid = s0.split(':', 1)[1].strip(); continue
            if r[0] is None: continue
            out.append({'org_id': 'LUX', 'period': 'July 2026', 'store': store, 'salesperson': g(r, 'Salesperson'),
                        'department': str(g(r, 'Department')).strip(), 'category': str(g(r, cat)).strip(),
                        'product_desc': str(g(r, 'Product Desc')).strip(),
                        'contract_type': str(g(r, 'Contract Type')).strip(),
                        'ext_price': sf(g(r, 'Ext Price')), 'gp': sf(g(r, 'GP')),
                        'voided': str(g(r, 'Voided')).strip(), 'trans_type': str(g(r, 'Trans Type')).strip(),
                        'trans_id': tid, 'trans_date': str(g(r, 'Trans Date Time'))[:10]})
        return out

    # minimal FakeClient mirroring what _exec_mtd reads (feed for the open month, empty raw_sales).
    class FakeTable:
        def __init__(self, rows): self._rows = rows; self._f = []
        def select(self, *a, **k): return self
        def eq(self, c, v): self._f.append((c, v)); return self
        def in_(self, c, v): return self
        def order(self, *a, **k): return self
        def range(self, *a, **k): return self
        def limit(self, *a, **k): return self
        def execute(self):
            rows = [r for r in self._rows if all(r.get(c) == v for c, v in self._f if c in r)]
            class _R: data = rows
            return _R()
    class FakeSchema:
        def __init__(self, data): self._data = data
        def table(self, name): return FakeTable(self._data.get(name, []))
        def rpc(self, *a, **k):
            class _R: data = []
            return _R()
    class FakeClient:
        def __init__(self, data): self._data = data
        def schema(self, *a, **k): return FakeSchema(self._data)
        def table(self, name): return FakeTable(self._data.get(name, []))
        def rpc(self, *a, **k):
            class _R: data = []
            return _R()

    for cat in ('Category', 'System Category'):
        feed = build_feed(cat)
        c = FakeClient({'daily_sales_feed': feed, 'raw_sales': [], 'store_mapping': [],
                        'report_definitions': [], 'accessory_config': [], 'flag_rules': [],
                        'gp_category_map': [], 'exec_metric_config': [], 'storeops.employees': [],
                        'targets': []})
        try:
            res = R._exec_mtd(c, 'LUX', 'July 2026')
            tot = res['by_location']['total']
            ta = tot['total_activation']; port = tot.get('port'); actv = tot.get('activation')
            print(f"  [{cat}] _exec_mtd total_activation={ta} activation(non-port)={actv} port={port} "
                  f"byod={tot.get('byod')} upgrade={tot.get('upgrade')}")
            # Engine-verified deltas (OLD -> NEW), distinct-txn:
            #   total_activation 21 -> 26  (+5 brand-new Port-with-IDV-only transactions)
            #   premium/Sales-Report activations 14 -> 19  (+5)
            #   port sub-split   1 -> 7   (+6: the 5 new + 1 txn that was already premium via an
            #                              'Activation' line but ALSO carries a Port-with-IDV line,
            #                              now recognized as a Port)
            #   activation(non-port) 13 -> 12  (-1: that 1 txn moves from plain-activation to Port)
            #   byod 6, upgrade 1  UNCHANGED
            check(f"[{cat}] total_activation == 26 (was 21; +5 Port-with-IDV-only txns)", ta == 26, f"got {ta}")
            check(f"[{cat}] premium/Sales-Report activations == 19 (was 14; +5)",
                  tot.get('activation') + tot.get('port') == 19, f"got {tot.get('activation')+tot.get('port')}")
            check(f"[{cat}] port == 7 (was 1; 5 new + 1 already-premium txn now recognized Port)", port == 7, f"got {port}")
            check(f"[{cat}] activation(non-port) == 12 (was 13; -1 reclassified to Port)", actv == 12, f"got {actv}")
            check(f"[{cat}] byod == 6 UNCHANGED", tot.get('byod') == 6, f"got {tot.get('byod')}")
            check(f"[{cat}] upgrade == 1 UNCHANGED", tot.get('upgrade') == 1, f"got {tot.get('upgrade')}")
        except Exception as e:
            check(f"[{cat}] _exec_mtd ran", False, f"raised {type(e).__name__}: {e}")
else:
    print(f"  (SKIP real-sample engine leg — sample not present at {SAMPLE})")

print(f"\n{'PASS' if FAIL == 0 else 'FAIL'}  {PASS}/{PASS + FAIL}")
sys.exit(1 if FAIL else 0)
