"""Proof harness for agent/commission/luxelink-sales-flow.

Drives the REAL router functions (_exec_mtd, _exec_line_match, _exec_act_class, _promote_all_due,
_promote_feed_to_raw_sales, _sales_rows_union) over an in-memory FakeClient + the REAL luxelink
Total-Wireless sample file. No DB, no network. Run: python3 backend/scratchpad/luxelink_sales_flow_proof.py
"""
import os, sys
from collections import Counter
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import app.modules.commcalc.router as R

PASS = 0; FAIL = 0
def check(name, cond, extra=""):
    global PASS, FAIL
    if cond: PASS += 1; print(f"  ok  {name}")
    else: FAIL += 1; print(f"FAIL  {name}   {extra}")

# ── in-memory fake supabase client ───────────────────────────────────────────────────────────────
class FakeResult:
    def __init__(self, data=None, count=None): self.data = data or []; self.count = count
class _RpcExec:
    def __init__(self, data): self._d = data
    def execute(self): return FakeResult(data=self._d)
class FakeQuery:
    def __init__(self, store, table):
        self.store = store; self.t = table; self.f = []; self.cnt = False
        self.op = 'select'; self.ins = None; self.rng = None
    def select(self, *a, **k):
        if k.get('count') == 'exact': self.cnt = True
        return self
    def eq(self, c, v): self.f.append(('eq', c, v)); return self
    def in_(self, c, v): self.f.append(('in', c, list(v))); return self
    def neq(self, c, v): self.f.append(('neq', c, v)); return self
    def limit(self, n): return self
    def range(self, a, b): self.rng = (a, b); return self
    def order(self, *a, **k): return self
    def delete(self): self.op = 'delete'; return self
    def insert(self, rows): self.op = 'insert'; self.ins = rows if isinstance(rows, list) else [rows]; return self
    def upsert(self, rows, **k): self.op = 'upsert'; self.ins = rows if isinstance(rows, list) else [rows]; return self
    def _m(self, r):
        for k, c, v in self.f:
            rv = r.get(c)
            if k == 'eq' and rv != v: return False
            if k == 'in' and rv not in v: return False
            if k == 'neq' and rv == v: return False
        return True
    def execute(self):
        rows = self.store.setdefault(self.t, [])
        if self.op == 'select':
            m = [r for r in rows if self._m(r)]
            if self.rng: a, b = self.rng; m = m[a:b + 1]
            if self.cnt: return FakeResult(data=m, count=len(m))
            return FakeResult(data=[dict(r) for r in m])
        if self.op == 'delete':
            self.store[self.t] = [r for r in rows if not self._m(r)]
            return FakeResult(data=[])
        if self.op in ('insert', 'upsert'):
            for r in self.ins: rows.append(dict(r))
            return FakeResult(data=list(self.ins))
        return FakeResult()
class FakeSchema:
    def __init__(self, store): self.store = store
    def table(self, t): return FakeQuery(self.store, t)
    def rpc(self, name, params):
        if name == 'sales_feed_orgs_for_period':
            pv = params.get('p_periods') or []
            c = Counter(r['org_id'] for r in self.store.get('daily_sales_feed', []) if r.get('period') in pv)
            return _RpcExec([{'org_id': k, 'feed_rows': v} for k, v in c.items()])
        raise Exception('no such rpc')
class FakeClient:
    def __init__(self, store): self.store = store
    def schema(self, s): return FakeSchema(self.store)

def new_client(store=None):
    store = store if store is not None else {}
    c = FakeClient(store)
    R.sb = lambda: c          # any internal sb() (e.g. _write_upload_trace) uses the fake too
    return c, store

# ════════════════════════════════════════════════════════════════════════════════════════════════
print("\n== A. EXEC MTD engine over the REAL luxelink sample file ==")
import openpyxl
SAMPLE = "/workspaces/commcalc/commcalc/My Sales Transaction Details Legacy New with all columns (3).xlsx"
wb = openpyxl.load_workbook(SAMPLE, data_only=True)
ws = wb[wb.sheetnames[0]]; xr = list(ws.iter_rows(values_only=True)); hdr = list(xr[0]); IDX = {h: i for i, h in enumerate(hdr)}
def g(r, n): v = r[IDX[n]]; return '' if v is None else v
def sf(x):
    try: return float(x)
    except Exception: return 0.0

def build_feed(cat_source):
    """cat_source: 'Category' or 'System Category' — mimics the ingest's category = Category OR System Cat."""
    store = None; tid = None; out = []
    for r in xr[1:]:
        s0 = str(r[0]).strip() if r[0] is not None else ''
        if s0.startswith('Store:'): store = s0.split(':', 1)[1].strip(); continue
        if s0.startswith('Trans ID:'): tid = s0.split(':', 1)[1].strip(); continue
        if r[0] is None: continue
        out.append({'org_id': 'LUX', 'period': 'July 2026', 'store': store, 'salesperson': g(r, 'Salesperson'),
                    'department': str(g(r, 'Department')).strip(), 'category': str(g(r, cat_source)).strip(),
                    'product_desc': str(g(r, 'Product Desc')).strip(), 'contract_type': str(g(r, 'Contract Type')).strip(),
                    'ext_price': sf(g(r, 'Ext Price')), 'gp': sf(g(r, 'GP')),
                    'voided': str(g(r, 'Voided')).strip(), 'trans_type': str(g(r, 'Trans Type')).strip(),
                    'trans_id': tid, 'trans_date': str(g(r, 'Trans Date Time'))[:10]})
    return out

def reference(feed):
    """Independent re-implementation of the bucket counts (excludes voided) — cross-checks the engine
    rather than trusting hardcoded numbers. The 330-row sample has 14 voided rows (3 with a contract type:
    1 Activation + 2 BYOD), which the exec report correctly EXCLUDES → 72 activations, not the raw 75."""
    e = {'total_activation': 0, 'activation': 0, 'port': 0, 'byod': 0, 'upgrade': 0,
         'total_phones': 0, 'bill_payment_qty': 0}
    for r in feed:
        if str(r['voided']).strip().lower() in ('yes', 'true', '1', 'y', 't', 'voided'):
            continue
        ct = r['contract_type'].lower(); cat = r['category'].lower(); dept = r['department'].lower()
        if ct:
            e['total_activation'] += 1
            e['upgrade' if 'upgrade' in ct else 'byod' if 'byod' in ct else 'port' if 'port' in ct else 'activation'] += 1
        if cat in ('cellphone', 'kittedbranded'):
            e['total_phones'] += 1
        if dept == 'rtr' or cat in ('rtr product', 'other carr. payments'):
            e['bill_payment_qty'] += 1
    return e

for cat_source in ('Category', 'System Category'):
    feed = build_feed(cat_source)
    c, store = new_client({'daily_sales_feed': feed, 'raw_sales': []})
    res = R._exec_mtd(c, 'LUX', 'July 2026')
    tot = res['by_location']['total']
    exp = reference(feed)
    for k, v in exp.items():
        check(f"[{cat_source}] {k} == {v} (void-excluded)", tot[k] == v, f"got {tot[k]}")
    check(f"[{cat_source}] Total Activation == parts",
          tot['total_activation'] == tot['activation'] + tot['port'] + tot['byod'] + tot['upgrade'])
    # both export variants (Category vs System Category) must give IDENTICAL bucket counts
    check(f"[{cat_source}] total_activation == 72 (14 voided rows, 3 with CT, correctly excluded)",
          tot['total_activation'] == 72, f"got {tot['total_activation']}")
    # location vs employee totals reconcile (same lines, different grouping)
    et = res['by_employee']['total']
    for k in ('total_activation', 'total_phones', 'bill_payment_qty', 'acc_sales', 'activation_fee', 'total_protect'):
        check(f"[{cat_source}] loc total == emp total ({k})", abs(tot[k] - et[k]) < 0.01, f"{tot[k]} vs {et[k]}")

print("\n== A2. formula checks (Trending / Conv / APB) on the sample ==")
c, store = new_client({'daily_sales_feed': build_feed('Category'), 'raw_sales': []})
res = R._exec_mtd(c, 'LUX', 'July 2026')
tr = res['trending']; tot = res['by_location']['total']
check("trending elapsed=14 (Jul 15 → yesterday=14)", tr['elapsed_days'] == 14, f"got {tr['elapsed_days']} (only true when run on 2026-07-15)")
check("trending days_in_month=31", tr['days_in_month'] == 31)
check("Conv == TotalAct / BillPayQty",
      abs(tot['conv'] - round(tot['total_activation'] / tot['bill_payment_qty'], 4)) < 1e-9)
check("APB == Acc.Sales / TotalAct",
      abs(tot['apb'] - round(tot['acc_sales'] / tot['total_activation'], 2)) < 1e-9)
# per-row: Trending Box == round(TotalAct * factor); Trending Acc == round(Acc * factor,2)
row0 = res['by_location']['rows'][0]
f = tr['factor']
check("per-row Trending Box == round(TotalAct*factor)", row0['trending_box'] == round(row0['total_activation'] * f))
check("per-row Trending Acc == round(Acc*factor,2)", abs(row0['trending_acc_sales'] - round(row0['acc_sales'] * f, 2)) < 0.01)
# totals row = SUM of per-row trending (matches how the spreadsheet totals reconcile: 793)
check("total Trending Box == sum(per-row trending box)",
      tot['trending_box'] == sum(r['trending_box'] for r in res['by_location']['rows']))

print("\n== A3. line-match + activation-split unit truth table ==")
ph = R._EXEC_METRIC_DEFAULTS['phones']['rules']
bp = R._EXEC_METRIC_DEFAULTS['bill_payment']['rules']
ac = R._EXEC_METRIC_DEFAULTS['accessory']['rules']
pr = R._EXEC_METRIC_DEFAULTS['protect']['rules']
ar = R._EXEC_METRIC_DEFAULTS['activation']['rules']
check("phones: CellPhone matches", R._exec_line_match(ph, '', 'cellphone', 'apple iphone'))
check("phones: KittedBranded matches", R._exec_line_match(ph, 'brandedhandset', 'kittedbranded', 'tcl tab'))
check("phones: accessory does NOT match", not R._exec_line_match(ph, '', 'accessory', 'case'))
check("bill_payment: dept rtr matches", R._exec_line_match(bp, 'rtr', 'other carr. payments', 'wallet'))
check("accessory: HandsetBranded case matches", R._exec_line_match(ac, 'brandedhandset', 'handsetbranded', 'case byod'))
check("protect: Device Protection matches", R._exec_line_match(pr, '', '', 'device protection'))
check("protect: Screen Protector EXCLUDED", not R._exec_line_match(pr, 'brandedhandset', 'handsetbranded', 'screen protectors byod'))
check("protect: RTR protection refill EXCLUDED (dept rtr)", not R._exec_line_match(pr, 'rtr', 'other carr. payments', 'total wireless protection rtr'))
check("act: 'Port with IDV' -> port", R._exec_act_class('Port with IDV', ar) == 'port')
check("act: 'BYOD Port' -> byod (priority)", R._exec_act_class('BYOD Port', ar) == 'byod')
check("act: 'Activation With IDV' -> activation", R._exec_act_class('Activation With IDV', ar) == 'activation')
check("act: 'Upgrade' -> upgrade", R._exec_act_class('Upgrade', ar) == 'upgrade')
check("act: blank -> None (not an activation line)", R._exec_act_class('', ar) is None)

print("\n== A4. voided lines excluded ==")
feed = build_feed('Category')
feed2 = [dict(x) for x in feed]
# void every activation line → total activation must drop to 0
for x in feed2:
    if x['contract_type']: x['voided'] = 'Yes'
c, store = new_client({'daily_sales_feed': feed2, 'raw_sales': []})
res2 = R._exec_mtd(c, 'LUX', 'July 2026')
check("voided activation lines excluded", res2['by_location']['total']['total_activation'] == 0,
      f"got {res2['by_location']['total']['total_activation']}")

print("\n== B. promotion is MONEY-SAFE: no double-count (trans_id in both tables counts once) ==")
# feed has trans A,B,C (July); raw_sales already has A (overlap) + D (monthly-only)
feed = [{'org_id': 'LUX', 'period': 'July 2026', 'trans_id': t, 'trans_date': '2026-07-05',
         'ext_price': 10.0, 'store': 's', 'salesperson': 'r', 'category': 'x'} for t in ('A', 'B', 'C')]
raw = [{'org_id': 'LUX', 'period': 'July 2026', 'trans_id': 'A', 'trans_date': '2026-07-01', 'ext_price': 99.0},
       {'org_id': 'LUX', 'period': 'July 2026', 'trans_id': 'D', 'trans_date': '2026-07-02', 'ext_price': 7.0}]
c, store = new_client({'daily_sales_feed': feed, 'raw_sales': [dict(x) for x in raw]})
summ = R._promote_feed_to_raw_sales(c, 'LUX', 'July 2026')
final = store['raw_sales']
tids = [r['trans_id'] for r in final]
check("result has A,B,C,D exactly (D kept, A not duplicated)", sorted(set(tids)) == ['A', 'B', 'C', 'D'])
check("A appears ONCE (no double-count across feed+raw)", tids.count('A') == 1, f"A count={tids.count('A')}")
check("A's row is the FEED row (feed wins the overlap)", next(r for r in final if r['trans_id'] == 'A')['ext_price'] == 10.0)
check("monthly-only D preserved with its value", next(r for r in final if r['trans_id'] == 'D')['ext_price'] == 7.0)

print("\n== B2. first-tenant promotion into EMPTY raw_sales (the fixed 500) ==")
feed = [{'org_id': 'NEW', 'period': 'July 2026', 'trans_id': t, 'trans_date': '2026-07-05',
         'ext_price': 5.0, 'store': 's', 'salesperson': 'r', 'category': 'c', 'feed_only_col': 'z'} for t in ('X', 'Y')]
c, store = new_client({'daily_sales_feed': feed, 'raw_sales': []})
summ = R._promote_feed_to_raw_sales(c, 'NEW', 'July 2026')
check("empty raw_sales promotes without raising", summ.get('written') == 2, f"summ={summ}")
check("promoted rows re-stamped org_id + canon period",
      all(r['org_id'] == 'NEW' and r['period'] == 'July 2026' for r in store['raw_sales']))

print("\n== C. _promote_all_due iterates EVERY org with feed rows (org-agnostic + self-healing) ==")
feed = ([{'org_id': 'LUX', 'period': 'July 2026', 'trans_id': f'L{i}', 'trans_date': '2026-07-05', 'ext_price': 1.0} for i in range(3)]
        + [{'org_id': 'HOUSE', 'period': 'July 2026', 'trans_id': f'H{i}', 'trans_date': '2026-07-05', 'ext_price': 1.0} for i in range(2)]
        + [{'org_id': 'OPT', 'period': 'July 2026', 'trans_id': 'O1', 'trans_date': '2026-07-05', 'ext_price': 1.0}])
c, store = new_client({'daily_sales_feed': feed, 'raw_sales': [],
                       # OPT opts out via report_definitions.auto=false for report_key='sales'
                       'report_definitions': [{'org_id': 'OPT', 'report_key': 'sales', 'auto': False}]})
out = R._promote_all_due(c, 'July 2026')
byorg = {d['org_id']: d for d in out['detail']}
check("all 3 feed orgs enumerated", out['orgs'] == 3, f"orgs={out['orgs']}")
check("LUX promoted (3 rows)", byorg['LUX'].get('written') == 3, f"{byorg.get('LUX')}")
check("HOUSE promoted (2 rows)", byorg['HOUSE'].get('written') == 2, f"{byorg.get('HOUSE')}")
check("OPT skipped (sales auto=false)", byorg['OPT'].get('skipped') == 'sales auto=false', f"{byorg.get('OPT')}")
check("raw_sales now holds LUX+HOUSE rows (5), OPT none",
      len([r for r in store['raw_sales'] if r['org_id'] in ('LUX', 'HOUSE')]) == 5
      and not any(r['org_id'] == 'OPT' for r in store['raw_sales']))
# idempotent: a second run must NOT duplicate (delete-then-insert on the same feed)
R._promote_all_due(c, 'July 2026')
check("idempotent: second run doesn't duplicate rows",
      len([r for r in store['raw_sales'] if r['org_id'] in ('LUX', 'HOUSE')]) == 5,
      f"got {len([r for r in store['raw_sales'] if r['org_id'] in ('LUX','HOUSE')])}")

print("\n== C2. _promote_all_due falls back to a scan when the RPC is absent ==")
class NoRpcSchema(FakeSchema):
    def rpc(self, name, params): raise Exception('function does not exist')
class NoRpcClient(FakeClient):
    def schema(self, s): return NoRpcSchema(self.store)
store = {'daily_sales_feed': [{'org_id': 'LUX', 'period': 'July 2026', 'trans_id': 'L1',
                               'trans_date': '2026-07-05', 'ext_price': 1.0}], 'raw_sales': []}
c = NoRpcClient(store); R.sb = lambda: c
out = R._promote_all_due(c, 'July 2026')
check("scan fallback still finds the org", out['orgs'] == 1 and out['written_orgs'] == 1, f"{out}")

print("\n== D. union display source shows the feed for the OPEN month w/ EMPTY raw_sales (the symptom) ==")
feed = build_feed('Category')
c, store = new_client({'daily_sales_feed': feed, 'raw_sales': []})
rows, meta = R._sales_rows_union(c, 'LUX', 'July 2026')
check("union returns feed rows when raw_sales empty (report not starved)", len(rows) == len(feed), f"{len(rows)} vs {len(feed)}")
check("union meta: feed led", meta['primary'] == 'daily_sales_feed' and meta['raw_rows'] == 0)

print(f"\n=================  {PASS} passed, {FAIL} failed  =================")
sys.exit(1 if FAIL else 0)
